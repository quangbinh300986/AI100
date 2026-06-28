"""
每日填报接口
提供每日填报的CRUD和审核操作API
"""

from datetime import date, datetime, timezone
from typing import Optional
import uuid
import httpx
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, BackgroundTasks
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.config import settings
from app.database import get_db, get_crm_db
from app.models.user import User, UserRole
from app.models.report import DailyReport, ReportDetail, ReportStatus, WeeklyReport, GroupWeeklyReport
from app.schemas.report import (
    DailyReportCreate,
    DailyReportUpdate,
    DailyReportResponse,
    ReportReviewRequest,
    ReportListResponse,
    WeeklyReportCreate,
    WeeklyReportUpdate,
    WeeklyReportResponse,
    WeeklyReportListResponse,
    WeeklyCrmSummaryListResponse,
    WeeklyCrmSummaryItem,
    GroupWeeklyReportSave,
    GroupWeeklyReportResponse,
)
from fastapi import Request
from app.api.deps import get_current_user, require_permission
from app.services.audit_service import log_action, to_dict

# 将本文件内所有的 require_roles 拦截器重映射到 reports 动态权限校验上，实现精细化数据库动态控制
def dynamic_require_roles(*roles):
    async def reports_permission_dependency(
        request: Request,
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
    ) -> User:
        if current_user.role == UserRole.ADMIN:
            return current_user
            
        path = request.url.path
        
        # 统一映射到 approve_report，只要勾选了填报审核大类，均允许审核操作
        perm = "approve_report"
        
        checker = require_permission(perm)
        return await checker(current_user, db)
        
    return reports_permission_dependency

require_roles = dynamic_require_roles

router = APIRouter(prefix="/reports", tags=["每日填报"])


@router.post("/upload", summary="上传填报图片附件")
async def upload_photo(
    file: UploadFile = File(..., description="填报照片图片"),
    current_user: User = Depends(get_current_user),
):
    """
    上传照片至 Supabase 的 photos 存储桶中，返回公开访问 URL
    """
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="只能上传图片文件")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="图片大小不能超过 10MB")

    # 生成唯一文件名
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{uuid.uuid4().hex}.{ext}"

    # 构造 Supabase Storage 对象的上传 URL
    upload_url = f"{settings.SUPABASE_URL}/storage/v1/object/photos/{filename}"

    headers = {
        "Authorization": f"Bearer {settings.SERVICE_ROLE_KEY}",
        "Content-Type": file.content_type
    }

    async with httpx.AsyncClient() as client:
        try:
            resp = await client.post(upload_url, content=content, headers=headers, timeout=10.0)
            if resp.status_code not in [200, 201]:
                raise HTTPException(status_code=500, detail=f"上传至 Supabase Storage 失败: {resp.text}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")

    # 获取公开 URL，支持公网穿透域名替换
    supabase_url = settings.SUPABASE_URL
    if getattr(settings, "EXTERNAL_SUPABASE_URL", None):
        supabase_url = settings.EXTERNAL_SUPABASE_URL
    public_url = f"{supabase_url.rstrip('/')}/storage/v1/object/public/photos/{filename}"

    return {"url": public_url, "filename": filename}



@router.post("", response_model=DailyReportResponse, status_code=201, summary="创建每日填报")
async def create_report(
    report_in: DailyReportCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    创建当日填报记录
    每人每天只能有一条填报记录（由数据库唯一约束保证）
    """
    # 检查是否已有当日填报
    existing = await db.execute(
        select(DailyReport).where(
            DailyReport.user_id == current_user.id,
            DailyReport.report_date == report_in.report_date,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"已存在{report_in.report_date}的填报记录",
        )

    # 统计明细，自动计算主表指标数值，不再使用前端输入的原字段手填值
    c_amount = 0.0
    c_count = 0
    h_actions = 0
    t_count = 0
    l_count = 0
    pl_count = 0

    from app.models.report import DetailType

    for detail_in in report_in.details:
        if detail_in.detail_type == DetailType.CONTRACT:
            c_amount += detail_in.amount or 0.0
            c_count += 1
        elif detail_in.detail_type == DetailType.HAPPINESS:
            h_actions += 1
        elif detail_in.detail_type == DetailType.TRIANGLE:
            t_count += 1
        elif detail_in.detail_type == DetailType.LEAD:
            if detail_in.lead_progress == "25%":
                l_count += 1
        elif detail_in.detail_type == DetailType.POTENTIAL_LEAD:
            pl_count += 1

    # 创建填报主记录
    report = DailyReport(
        user_id=current_user.id,
        report_date=report_in.report_date,
        contract_amount=c_amount,
        contract_count=c_count,
        happiness_actions=h_actions,
        triangle_count=t_count,
        leads_count=l_count,
        potential_leads_count=pl_count,
        work_summary=report_in.work_summary,
        work_reflection=report_in.work_reflection,
        next_day_plan=report_in.next_day_plan,
        standup_notes=report_in.standup_notes,
        status=ReportStatus.DRAFT,
    )
    db.add(report)
    await db.flush()

    # 创建明细记录
    for detail_in in report_in.details:
        detail = ReportDetail(
            report_id=report.id,
            detail_type=detail_in.detail_type,
            customer_name=detail_in.customer_name,
            amount=detail_in.amount,
            lead_progress=detail_in.lead_progress,
            crm_opportunity_id=detail_in.crm_opportunity_id,
            happiness_level=detail_in.happiness_level,
            happiness_standard_id=detail_in.happiness_standard_id,
            description=detail_in.description,
            attachment_urls=detail_in.attachment_urls,
            partner_user_id=detail_in.partner_user_id,
        )
        db.add(detail)

    await db.flush()
    await db.refresh(report)

    # 加载明细及搭档姓名
    result = await db.execute(
        select(DailyReport)
        .options(selectinload(DailyReport.details).selectinload(ReportDetail.partner_user))
        .where(DailyReport.id == report.id)
    )
    loaded_report = result.scalar_one()
    for d in loaded_report.details:
        d.partner_name = d.partner_user.name if d.partner_user else None
    return loaded_report


@router.get("", response_model=ReportListResponse, summary="获取填报列表")
async def list_reports(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    report_date: date | None = Query(None, description="按日期筛选"),
    user_id: int | None = Query(None, description="按用户筛选"),
    team_id: int | None = Query(None, description="按战队筛选"),
    status_filter: str | None = Query(None, alias="status", description="按状态筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    获取填报列表，支持分页和筛选
    队员只能看自己的，战队长只能看本战队成员的，超级管理员/目标官/数字专员可看所有
    """
    query = select(DailyReport).options(selectinload(DailyReport.details))

    # 1. 角色权限过滤
    if current_user.role in [UserRole.STAFF.value, UserRole.MARKETING_STAFF.value, UserRole.TECH_MARKETING.value]:
        # 队员（普通、营销、技术营销）强制只看自己的日报
        query = query.where(DailyReport.user_id == current_user.id)
    elif current_user.role == UserRole.TEAM_LEADER.value:
        # 战队长强制只看自己战队成员的日报
        if current_user.team_id is None:
            # 若战队长未分配战队，则无法看到任何数据
            query = query.where(DailyReport.id == -1)
        else:
            # JOIN User 表，强制限定 team_id
            query = query.join(User, DailyReport.user_id == User.id).where(User.team_id == current_user.team_id)
            # 如果战队长传入了特定 user_id 筛选，确保该用户在同一战队内
            if user_id is not None:
                query = query.where(DailyReport.user_id == user_id)
    else:
        # 超级管理员、目标官、数字专员：支持灵活按战队或个人过滤
        if team_id is not None:
            query = query.join(User, DailyReport.user_id == User.id).where(User.team_id == team_id)
        if user_id is not None:
            query = query.where(DailyReport.user_id == user_id)

    # 2. 日期筛选
    if report_date is not None:
        query = query.where(DailyReport.report_date == report_date)

    # 3. 状态筛选
    if status_filter is not None:
        query = query.where(DailyReport.status == status_filter)

    # 4. 排序：按日期降序
    query = query.order_by(DailyReport.report_date.desc())

    # 5. 正确计算过滤后的总条数
    count_query = select(func.count()).select_from(query.subquery())
    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0

    # 6. 分页
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    reports = result.scalars().all()

    return ReportListResponse(total=total, items=reports)



@router.put("/{report_id}", response_model=DailyReportResponse, summary="更新填报")
async def update_report(
    report_id: int,
    report_in: DailyReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    更新填报记录
    只有草稿和被驳回状态的记录可以编辑
    """
    result = await db.execute(
        select(DailyReport)
        .options(selectinload(DailyReport.details))
        .where(DailyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=404, detail="填报记录不存在")

    # 权限检查
    if report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能编辑自己的填报记录")

    # 状态检查
    if report.status not in [ReportStatus.DRAFT.value, ReportStatus.REJECTED.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="当前状态不允许编辑",
        )

    # 更新非空字段
    update_data = report_in.model_dump(exclude_unset=True, exclude={"details"})
    for field, value in update_data.items():
        setattr(report, field, value)

    # 更新明细（如果提供了）
    if report_in.details is not None:
        # 删除旧明细
        for detail in report.details:
            await db.delete(detail)
            
        c_amount = 0.0
        c_count = 0
        h_actions = 0
        t_count = 0
        l_count = 0
        pl_count = 0
        
        # 创建新明细并重新计算数值
        for detail_in in report_in.details:
            detail = ReportDetail(
                report_id=report.id,
                detail_type=detail_in.detail_type,
                customer_name=detail_in.customer_name,
                amount=detail_in.amount,
                lead_progress=detail_in.lead_progress,
                crm_opportunity_id=detail_in.crm_opportunity_id,
                happiness_level=detail_in.happiness_level,
                happiness_standard_id=detail_in.happiness_standard_id,
                description=detail_in.description,
                attachment_urls=detail_in.attachment_urls,
                partner_user_id=detail_in.partner_user_id,
            )
            db.add(detail)
            
            if detail_in.detail_type == DetailType.CONTRACT:
                c_amount += detail_in.amount or 0.0
                c_count += 1
            elif detail_in.detail_type == DetailType.HAPPINESS:
                h_actions += 1
            elif detail_in.detail_type == DetailType.TRIANGLE:
                t_count += 1
            elif detail_in.detail_type == DetailType.LEAD:
                if detail_in.lead_progress == "25%":
                    l_count += 1
            elif detail_in.detail_type == DetailType.POTENTIAL_LEAD:
                pl_count += 1
                
        report.contract_amount = c_amount
        report.contract_count = c_count
        report.happiness_actions = h_actions
        report.triangle_count = t_count
        report.leads_count = l_count
        report.potential_leads_count = pl_count

    db.add(report)
    await db.flush()
    await db.refresh(report)

    # 重新加载明细及搭档姓名
    result = await db.execute(
        select(DailyReport)
        .options(selectinload(DailyReport.details).selectinload(ReportDetail.partner_user))
        .where(DailyReport.id == report.id)
    )
    loaded_report = result.scalar_one()
    for d in loaded_report.details:
        d.partner_name = d.partner_user.name if d.partner_user else None
    return loaded_report


@router.post("/{report_id}/submit", summary="提交填报")
async def submit_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """将草稿状态的填报提交并默认自动审核通过"""
    result = await db.execute(
        select(DailyReport).where(DailyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=404, detail="填报记录不存在")

    if report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能提交自己的填报记录")

    if report.status not in [ReportStatus.DRAFT.value, ReportStatus.REJECTED.value]:
        raise HTTPException(status_code=400, detail="当前状态不允许提交")

    # 现阶段默认提交直接审核通过
    report.status = ReportStatus.REVIEWED
    report.submitted_at = datetime.now(timezone.utc)
    report.reviewed_at = datetime.now(timezone.utc)
    report.reviewer_id = None
    db.add(report)
    await db.flush()

    # 记录审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="UPDATE",
        target_module="report",
        target_id=report.id,
        description=f"提交日报并自动审核通过：日期 {report.report_date}",
        before_state=None,
        after_state=to_dict(report)
    )

    # 自动通过后广播通知大屏刷新
    try:
        from app.services.websocket import ws_manager
        await ws_manager.broadcast({"type": "update", "event": "report_approved"})
    except Exception as e:
        # 记录日志，但不影响主逻辑返回
        import logging
        logging.getLogger("battle100").error(f"默认通过大屏 WebSocket 广播失败: {e}")

    return {"message": "填报已提交并默认通过"}


@router.post("/{report_id}/review", summary="审核填报")
async def review_report(
    report_id: int,
    review: ReportReviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.ADMIN, UserRole.TARGET_OFFICER, UserRole.TEAM_LEADER)
    ),
):
    """
    审核填报记录
    仅管理员、目标官、战队长可操作。其中战队长只能审核本战队成员。
    """
    result = await db.execute(
        select(DailyReport).where(DailyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=404, detail="填报记录不存在")

    # 战队长防越权审查校验
    if current_user.role == UserRole.TEAM_LEADER.value:
        report_user_res = await db.execute(
            select(User.team_id).where(User.id == report.user_id)
        )
        report_user_team_id = report_user_res.scalar()
        if current_user.team_id is None or report_user_team_id != current_user.team_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="权限不足，您无权审核其他战队成员的日报"
            )

    if report.status != ReportStatus.SUBMITTED.value:
        raise HTTPException(status_code=400, detail="只能审核已提交的填报")

    # 状态修改前备份
    before_state = to_dict(report)

    if review.action == "approved":
        report.status = ReportStatus.REVIEWED
    elif review.action == "rejected":
        report.status = ReportStatus.REJECTED
    else:
        raise HTTPException(status_code=400, detail="无效的审核动作")

    report.reviewer_id = current_user.id
    report.reviewed_at = datetime.now(timezone.utc)
    db.add(report)
    await db.flush()

    # 记录审计日志
    action_cn = "审核通过" if review.action == "approved" else "审核驳回"
    await log_action(
        db=db,
        user=current_user,
        action_type="UPDATE",
        target_module="report",
        target_id=report.id,
        description=f"审核日报：日期 {report.report_date}，结果：{action_cn}。备注：{review.comment or '无'}",
        before_state=before_state,
        after_state=to_dict(report),
    )

    # 审核通过后广播通知大屏刷新
    if review.action == "approved":
        try:
            from app.services.websocket import ws_manager
            await ws_manager.broadcast({"type": "update", "event": "report_approved"})
        except Exception as e:
            # 记录日志，但不影响主逻辑返回
            import logging
            logging.getLogger("battle100").error(f"大屏 WebSocket 广播失败: {e}")

    return {"message": f"填报已{review.action}"}


# ==========================================
#              周复盘填报相关 API
# ==========================================

@router.get("/weekly/mine", response_model=WeeklyReportResponse, summary="获取我的周复盘填报")
async def get_my_weekly_report(
    start_date: date = Query(..., description="周开始日期(周一)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """根据开始日期查询当前登录用户的周复盘数据"""
    stmt = select(WeeklyReport).where(
        WeeklyReport.user_id == current_user.id,
        WeeklyReport.start_date == start_date
    )
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    
    if not report:
        raise HTTPException(status_code=404, detail="未找到该周的周报")
        
    # 补全用户名
    report.user_name = current_user.name
    return report


@router.post("/weekly", response_model=WeeklyReportResponse, summary="保存或提交周复盘填报")
async def save_weekly_report(
    report_in: WeeklyReportCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建或更新周复盘填报（支持保存草稿或直接提交）"""
    # 检查是否已存在记录
    stmt = select(WeeklyReport).where(
        WeeklyReport.user_id == current_user.id,
        WeeklyReport.start_date == report_in.start_date
    )
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    
    status_val = report_in.status if report_in.status else "draft"
    submitted_at_val = datetime.now(timezone.utc) if status_val == "submitted" else None
    
    if report:
        # 更新现有记录
        report.delivery_plan = report_in.delivery_plan
        report.sales_plan = report_in.sales_plan
        report.delivery_actual = report_in.delivery_actual
        report.sales_actual = report_in.sales_actual
        report.delivery_rate = report_in.delivery_rate
        report.sales_rate = report_in.sales_rate
        report.delivery_highlights = report_in.delivery_highlights
        report.sales_highlights = report_in.sales_highlights
        report.delivery_blockers = report_in.delivery_blockers
        report.sales_blockers = report_in.sales_blockers
        report.delivery_support = report_in.delivery_support
        report.sales_support = report_in.sales_support
        report.next_delivery_plan = report_in.next_delivery_plan
        report.next_sales_plan = report_in.next_sales_plan
        report.status = status_val
        if submitted_at_val:
            report.submitted_at = submitted_at_val
    else:
        # 创建新记录
        report = WeeklyReport(
            user_id=current_user.id,
            start_date=report_in.start_date,
            end_date=report_in.end_date,
            delivery_plan=report_in.delivery_plan,
            sales_plan=report_in.sales_plan,
            delivery_actual=report_in.delivery_actual,
            sales_actual=report_in.sales_actual,
            delivery_rate=report_in.delivery_rate,
            sales_rate=report_in.sales_rate,
            delivery_highlights=report_in.delivery_highlights,
            sales_highlights=report_in.sales_highlights,
            delivery_blockers=report_in.delivery_blockers,
            sales_blockers=report_in.sales_blockers,
            delivery_support=report_in.delivery_support,
            sales_support=report_in.sales_support,
            next_delivery_plan=report_in.next_delivery_plan,
            next_sales_plan=report_in.next_sales_plan,
            status=status_val,
            submitted_at=submitted_at_val
        )
        db.add(report)
        
    await db.flush()
    await db.commit()
    await db.refresh(report)
    
    # 补充用户名并记录日志
    report.user_name = current_user.name
    
    action_type_str = "SUBMIT" if status_val == "submitted" else "SAVE_DRAFT"
    await log_action(
        db=db,
        user=current_user,
        action_type=action_type_str,
        target_module="weekly_report",
        target_id=report.id,
        description=f"{'提交' if status_val == 'submitted' else '暂存'}周复盘，范围：{report.start_date} ~ {report.end_date}",
        before_state=None,
        after_state=to_dict(report)
    )
    
    if status_val == "submitted":
        # 取消发送机器人的周报卡片，改为在后台自动向钉钉工作日志提报并由钉钉系统自动投递官方日志卡片（解决大群/专属群消息重复问题）
        background_tasks.add_task(auto_sync_weekly_report_to_dingtalk_task, report.id, current_user.id)
        
    return report


@router.get("/weekly/auto-extract", summary="自动从播报系统提取周报实际完成内容")
async def extract_weekly_broadcasts(
    start_date: date = Query(..., description="周开始日期(周一)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    根据 Excel 附件 2 的要求，自动检索并提取该用户在对应周内自己填报的，
    以及该用户是联动人（提报明细或文字正则匹配）的五种核心动作播报信息。
    最后整理成排版优美的实际完成文本，直接提供给前端快速回填。
    """
    from datetime import timedelta, time
    import re
    from app.models.broadcast import BroadcastEvent
    from app.models.report import ReportDetail, DailyReport
    
    end_date = start_date + timedelta(days=6)
    
    # 转换为当周 UTC 时间边界
    start_dt = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
    
    # 1. 检索当前用户自己直接创建的播报
    stmt_self = select(BroadcastEvent).where(
        BroadcastEvent.user_id == current_user.id,
        BroadcastEvent.event_time >= start_dt,
        BroadcastEvent.event_time <= end_dt,
        BroadcastEvent.event_type.in_(["lead_25", "lead_75", "contract_signed", "triangle", "happiness", "potential_lead", "marketing_report", "middle_office_report", "happiness_committee"]),
        BroadcastEvent.is_deleted == False
    )
    res_self = await db.execute(stmt_self)
    broadcasts_self = res_self.scalars().all()
    
    # 2. 检索该用户在日报明细中作为联动搭档关联的播报事件
    stmt_linked = select(BroadcastEvent).join(
        ReportDetail,
        ReportDetail.description.like(func.concat('%[broadcast_id:', BroadcastEvent.id, ']%'))
    ).join(
        DailyReport,
        DailyReport.id == ReportDetail.report_id
    ).where(
        DailyReport.user_id == current_user.id,
        BroadcastEvent.event_time >= start_dt,
        BroadcastEvent.event_time <= end_dt,
        BroadcastEvent.event_type.in_(["lead_25", "lead_75", "contract_signed", "triangle", "happiness", "potential_lead", "marketing_report", "middle_office_report", "happiness_committee"]),
        BroadcastEvent.is_deleted == False
    )
    res_linked = await db.execute(stmt_linked)
    broadcasts_linked = res_linked.scalars().all()
    
    # 3. 正则兜底：查找包含“与联动人(姓名)”、“营销人员(姓名)”或“协助人(姓名)”并包含当前用户的当周所有播报
    stmt_all = select(BroadcastEvent).where(
        BroadcastEvent.event_time >= start_dt,
        BroadcastEvent.event_time <= end_dt,
        BroadcastEvent.event_type.in_(["lead_25", "lead_75", "contract_signed", "triangle", "happiness", "potential_lead", "marketing_report", "middle_office_report", "happiness_committee"]),
        BroadcastEvent.is_deleted == False
    )
    res_all = await db.execute(stmt_all)
    broadcasts_all = res_all.scalars().all()
    
    broadcasts_regex = []
    user_name = current_user.name
    for ev in broadcasts_all:
        content = ev.content or ""
        m1 = re.search(r"与联动人\((.*?)\)", content)
        m2 = re.search(r"营销人员\((.*?)\)", content)
        m3 = re.search(r"协助人[：:]\s*(.*?)(\n|$)", content)
        
        is_partner = False
        if m1:
            partners = [n.strip() for n in re.split(r'[，,、\s]', m1.group(1)) if n.strip()]
            if user_name in partners:
                is_partner = True
        if m2:
            partners = [n.strip() for n in re.split(r'[，,、\s]', m2.group(1)) if n.strip()]
            if user_name in partners:
                is_partner = True
        if m3:
            clean_names = m3.group(1).replace('*', '').strip()
            partners = [n.strip() for n in re.split(r'[，,、\s]', clean_names) if n.strip()]
            if user_name in partners:
                is_partner = True
                
        if is_partner:
            broadcasts_regex.append(ev)
            
    # 4. 合并去重并按发生时间升序排序
    event_dict = {}
    for ev in broadcasts_self + broadcasts_linked + broadcasts_regex:
        event_dict[ev.id] = ev
        
    all_events = list(event_dict.values())
    all_events.sort(key=lambda x: x.event_time if x.event_time else x.created_at)
    
    # 5. 统一合并排版为单个完成文本列表
    content_list = []
    for idx, ev in enumerate(all_events, 1):
        content_clean = ev.content or ""
        # 清除类似 [broadcast_id:123] 标记
        content_clean = re.sub(r"\s*\[broadcast_id:\d+\]", "", content_clean).strip()
        # 清除特定的口号文字（包含可能的前后换行、空格和常用标点逗号，以防排版零碎）
        content_clean = re.sub(r"奋战一百天[，,，]?\s*亮剑破六千[！!！]?\s*", "", content_clean)
        content_clean = re.sub(r"[，,，\s\n]*为客户幸福而奋斗[，,，]?\s*赢战百日[！!！]?", "", content_clean)
        content_clean = re.sub(r"[，,，\s\n]*赢战百日[！!！]?", "", content_clean)
        content_clean = content_clean.strip()
        
        prefix = ""
        if ev.event_type == "lead_75":
            prefix = "【中标】"
        elif ev.event_type == "lead_25":
            prefix = "【有效线索】"
        elif ev.event_type == "potential_lead":
            prefix = "【潜在线索确定】"
        elif ev.event_type == "contract_signed":
            prefix = "【合同签订】"
        elif ev.event_type == "triangle":
            prefix = " Barb 【铁三角联动】" if "奋战一百天" in content_clean else "【铁三角联动】"  # 保留可能的特定战役特征，以下做标准替换
            prefix = "【铁三角联动】"
        elif ev.event_type == "happiness":
            prefix = "【幸福动作】"
        elif ev.event_type == "marketing_report":
            prefix = "【营销内部播报】"
        elif ev.event_type == "middle_office_report":
            prefix = "【中台播报】"
        elif ev.event_type == "happiness_committee":
            prefix = "【中台幸福委播报】"
            
        content_list.append(f"{idx}. {prefix}{content_clean}")
        
    formatted_text = "\n".join(content_list) if content_list else "1. 无相关播报数据"
    
    # 导入所需的用户岗位和角色枚举
    from app.models.user import PositionType, UserRole
    
    # 判断当前用户是否为营销岗或目标官
    is_marketing = (
        current_user.position_type == PositionType.MARKETING 
        or current_user.role in [UserRole.TARGET_OFFICER, UserRole.MARKETING_STAFF, UserRole.TECH_MARKETING]
    )
    
    # 根据身份分岗归入对应的字段，不需填写的字段置空
    if is_marketing:
        return {
            "delivery_actual": "",
            "sales_actual": formatted_text
        }
    else:
        return {
            "delivery_actual": formatted_text,
            "sales_actual": ""
        }


def sync_extract_crm_data(real_name: str, start_date_val: date, is_marketing: bool) -> dict:
    from datetime import timedelta
    from sqlalchemy import text
    
    monday = start_date_val
    sunday = start_date_val + timedelta(days=6)
    
    # 格式化日期参数
    start_date_str = monday.strftime('%Y-%m-%d')
    end_date_str = sunday.strftime('%Y-%m-%d 23:59:59')
    
    result = {
        "delivery_actual": "",
        "sales_actual": "",
        "delivery_rate": "",
        "sales_rate": "",
        "delivery_highlights": "",
        "sales_highlights": "",
        "delivery_blockers": "",
        "sales_blockers": "",
        "delivery_support": "",
        "sales_support": "",
        "next_delivery_plan": "",
        "next_sales_plan": "",
        "crm_active_projects": "",
        "crm_milestone_tasks": "",
        "crm_suspended_projects": "",
        "crm_no_contract_warning": "",
        "crm_unbilled_warning": "",
        "crm_unreceived_warning": "",
        "crm_health_diagnosis": ""
    }
    
    try:
        from app.database import get_crm_db
        with get_crm_db() as conn:
            # 1. 统计个人当周产值与回款（排除重置虚高账期影响）
            month_start_date = monday.replace(day=1)
            prev_month_start_date = (month_start_date - timedelta(days=15)).replace(day=1)
            
            prod_sql = text("""
                SELECT COALESCE(SUM(dp.money), 0) as total_prod
                FROM dashboard_production dp
                JOIN project p ON dp.project_id = p.id
                WHERE p.project_manager = :real_name
                  AND dp.createDate BETWEEN :start AND :end
                  AND dp.account_date IN (:prev_month_start, :month_start)
                  AND dp.isDel = '0'
            """)
            prod_val = conn.execute(prod_sql, {
                "real_name": real_name,
                "start": start_date_str + " 00:00:00",
                "end": end_date_str,
                "prev_month_start": prev_month_start_date.strftime('%Y-%m-%d'),
                "month_start": month_start_date.strftime('%Y-%m-%d')
            }).scalar() or 0.0
            personal_production = float(prod_val) / 10000.0 # 元转万元
            
            recv_sql = text("""
                SELECT COALESCE(SUM(r.receive_money), 0) as total_recv
                FROM zdcrm_contract_receive_money_view r
                INNER JOIN contract c ON r.contract_id = c.id
                WHERE (c.signer = :real_name OR c.contract_head_user = :real_name)
                  AND r.create_date BETWEEN :start_date AND :end_date
            """)
            recv_val = conn.execute(recv_sql, {
                "real_name": real_name,
                "start_date": start_date_str,
                "end_date": end_date_str
            }).scalar() or 0.0
            personal_receive = float(recv_val) # 万元
            
            # 2. 统计个人名下正在实施项目并诊断饱和度预警
            active_projects_sql = text("""
                SELECT project_name, project_progress, project_status
                FROM project
                WHERE project_manager = :real_name
                  AND project_progress < 100.0
                  AND (project_status IS NULL OR (project_status != '已归档' AND project_status != '已结项' AND project_status != '3'))
            """)
            active_projects = conn.execute(active_projects_sql, {"real_name": real_name}).mappings().all()
            
            personal_warnings = []
            active_count = len(active_projects)
            
            if active_count == 0 and not is_marketing:
                personal_warnings.append("🚨 红色警报：您目前名下无任何活跃正在实施的交付项目，需立即核实饱和度并协调新项目分配！")
            else:
                # 检查项目本周是否进度停滞（无任何异动）
                p_change_sql = text("""
                    SELECT COUNT(*)
                    FROM dashboard_production dp
                    JOIN project p ON dp.project_id = p.id
                    WHERE p.project_manager = :real_name
                      AND dp.createDate BETWEEN :start AND :end
                      AND dp.account_date IN (:prev_month_start, :month_start)
                      AND dp.isDel = '0'
                """)
                change_count = conn.execute(p_change_sql, {
                    "real_name": real_name,
                    "start": start_date_str + " 00:00:00",
                    "end": end_date_str,
                    "prev_month_start": prev_month_start_date.strftime('%Y-%m-%d'),
                    "month_start": month_start_date.strftime('%Y-%m-%d')
                }).scalar() or 0
                
                if change_count == 0 and not is_marketing:
                    personal_warnings.append("⚠️ 黄色预警：名下正在实施项目本周进度停滞（无任何进度条推进记录），请在下方补充卡点或原因说明！")
                
                # 检查空仓风险
                all_near_complete = True
                for ap in active_projects:
                    progress_val = ap['project_progress']
                    try:
                        progress_val_float = float(progress_val) if progress_val is not None else 0.0
                    except Exception:
                        progress_val_float = 0.0
                    if progress_val_float < 90.0:
                        all_near_complete = False
                        break
                
                if active_count <= 2 and all_near_complete and not is_marketing:
                    personal_warnings.append(f"💡 风险提示：目前仅有 {active_count} 个正在实施项目且进度均已接近完成（当前进度≥90%），面临项目断档空仓风险，请尽快联系巴长安排新项目储备！")

            # 业绩快照与预警文本前缀准备
            perf_snapshot = (
                f"【📊 CRM 本周业绩快照】：累计确认产值 {personal_production:.2f} 万元，实际到账回款 {personal_receive:.2f} 万元。\n\n"
            )
            warning_text = ""
            if personal_warnings:
                warning_text = "【🚨 个人工作饱和度与项目健康度诊断】：\n" + "\n".join([f"  * {w}" for w in personal_warnings]) + "\n\n"

            # 3. 提取计划达成率说明 (根据 responsible_person_alias 匹配当前人在当月的目标)
            target_sql = text("""
                SELECT 
                    SUM(new_sign_target_amount) as target_sign, 
                    SUM(new_sign_actual_amount) as actual_sign,
                    SUM(receive_target_amount) as target_receive, 
                    SUM(receive_actual_amount) as actual_receive
                FROM zdcrm_target_plan_management
                WHERE (responsible_person_alias = :real_name OR create_by = :real_name)
                  AND year = :year AND month = :month
                  AND is_del = '0'
            """)
            target_rows = conn.execute(target_sql, {
                "real_name": real_name,
                "year": monday.year,
                "month": monday.month
            }).mappings().all()
            
            rate_desc = ""
            if target_rows and target_rows[0]["target_sign"] is not None:
                row = target_rows[0]
                t_sign = float(row["target_sign"])
                a_sign = float(row["actual_sign"]) if row["actual_sign"] else 0.0
                t_recv = float(row["target_receive"])
                a_recv = float(row["actual_receive"]) if row["actual_receive"] else 0.0
                
                sign_rate = f"{(a_sign / t_sign * 100):.1f}%" if t_sign > 0 else "—"
                recv_rate = f"{(a_recv / t_recv * 100):.1f}%" if t_recv > 0 else "—"
                
                if is_marketing:
                    rate_desc = f"月度销售新签达成率：{sign_rate} (实际 {a_sign:.1f}万 / 目标 {t_sign:.1f}万)；月度回款达成率：{recv_rate} (实际 {a_recv:.1f}万 / 目标 {t_recv:.1f}万)"
                else:
                    rate_desc = f"本月累计新签达成：{sign_rate}，回款达成：{recv_rate}"
            
            if is_marketing:
                result["sales_rate"] = rate_desc or "月度新签与回款指标正在统计中"
                result["delivery_rate"] = ""
            else:
                result["delivery_rate"] = rate_desc or "月度指标正在统计中"
                result["sales_rate"] = ""
                
            if is_marketing:
                # 2. 营销岗：新签合同 actual
                contract_sql = text("""
                    SELECT contract_name, contract_money 
                    FROM contract 
                    WHERE signer = :real_name 
                      AND signing_date BETWEEN :start_date AND :end_date 
                      AND (is_suspension IS NULL OR is_suspension != '1')
                """)
                contracts = conn.execute(contract_sql, {
                    "real_name": real_name,
                    "start_date": start_date_str,
                    "end_date": end_date_str
                }).mappings().all()
                
                sales_list = []
                c_idx = 1
                if contracts:
                    sales_list.append("本周正式签约合同项目如下：")
                    for c in contracts:
                        # 原始单位为“元”，需除以10000转换为“万元”
                        c_money_val = float(c['contract_money']) / 10000.0 if c['contract_money'] is not None else 0.0
                        sales_list.append(f"  {c_idx}) 【{c['contract_name']}】签署成功，金额：{c_money_val:.2f} 万元")
                        c_idx += 1
                
                # 3. 营销岗：实际回款 actual
                receive_sql = text("""
                    SELECT r.contract_name, r.receive_money, r.receive_date 
                    FROM zdcrm_contract_receive_money_view r
                    INNER JOIN contract c ON r.contract_id = c.id
                    WHERE (c.signer = :real_name OR c.contract_head_user = :real_name)
                      AND r.create_date BETWEEN :start_date AND :end_date
                """)
                receives = conn.execute(receive_sql, {
                    "real_name": real_name,
                    "start_date": start_date_str,
                    "end_date": end_date_str
                }).mappings().all()
                
                if receives:
                    if not sales_list:
                        sales_list.append("本周到账回款明细：")
                    else:
                        sales_list.append("\n本周到账回款明细：")
                    for r in receives:
                        sales_list.append(f"  {c_idx}) 【{r['contract_name']}】收到回款金额：{float(r['receive_money'] or 0):.2f} 万元，到账日期：{r['receive_date'].strftime('%m-%d') if r['receive_date'] else '—'}")
                        c_idx += 1
                
                # 4. 营销岗：客户跟进拜访 actual
                visit_sql = text("""
                    SELECT customer_name, remark, create_time 
                    FROM zdcrm_visit_customer_record 
                    WHERE (create_by = :real_name OR update_by = :real_name)
                      AND create_time BETWEEN :start_date AND :end_date
                      AND is_del = '0'
                    ORDER BY create_time ASC
                """)
                visits = conn.execute(visit_sql, {
                    "real_name": real_name,
                    "start_date": start_date_str,
                    "end_date": end_date_str
                }).mappings().all()
                
                if visits:
                    if not sales_list:
                        sales_list.append("本周拜访/跟进客户动作明细：")
                    else:
                        sales_list.append("\n本周拜访/跟进客户动作明细：")
                    for v in visits:
                        t_str = v['create_time'].strftime('%m-%d') if v['create_time'] else '—'
                        sales_list.append(f"  {c_idx}) 【{t_str}】对接拜访【{v['customer_name']}】(工作记录：{v['remark'] or '未填'})")
                        c_idx += 1
                
                formatted_sales = perf_snapshot + ("\n".join(sales_list) if sales_list else "1. 本周暂无相关的合同新签、到账回款与客户拜访登记。")
                result["sales_actual"] = formatted_sales
                result["delivery_actual"] = ""
                
                # 5. 亮点与卡点智能提取
                highlight_list = []
                blocker_list = []
                
                # 如果有大额签约
                h_idx = 1
                for c in contracts:
                    # 原始单位为“元”，需除以10000转换为“万元”
                    c_money_val = float(c['contract_money']) / 10000.0 if c['contract_money'] is not None else 0.0
                    if c_money_val >= 50.0:
                        highlight_list.append(f"  {h_idx}) 成功签订大额合同：【{c['contract_name']}】（金额：{c_money_val:.2f}万元）")
                        h_idx += 1
                
                if len(visits) >= 3:
                    highlight_list.append(f"  {h_idx}) 本周客户对接频次较高，累计完成 {len(visits)} 次客户拜访与商务洽谈")
                    h_idx += 1
                
                # 卡点检查
                terminated_sql = text("""
                    SELECT remark, action_strategy 
                    FROM zdcrm_target_plan_management
                    WHERE (responsible_person_alias = :real_name OR create_by = :real_name)
                      AND is_terminated = '1' AND is_del = '0'
                      AND year = :year AND month = :month
                """)
                term_projects = conn.execute(terminated_sql, {
                    "real_name": real_name,
                    "year": monday.year,
                    "month": monday.month
                }).mappings().all()
                
                b_idx = 1
                if term_projects:
                    for tp in term_projects:
                        blocker_list.append(f"  {b_idx}) CRM 中标注的中止/预警项目跟进阻碍（备注：{tp['remark'] or '无'}，应对策略：{tp['action_strategy'] or '暂无'}）")
                        b_idx += 1
                
                result["sales_highlights"] = "\n".join(highlight_list) if highlight_list else "1. 本周销售签约及商务拓展平稳推进。"
                result["sales_blockers"] = warning_text + ("\n".join(blocker_list) if blocker_list else "1. 目前名下意向商机及收款合同暂无重大异常阻碍。")
                result["delivery_highlights"] = ""
                result["delivery_blockers"] = ""
                
            else:
                # 6. 交付及其他岗：负责的项目与进度 update（筛选最近一个月内更新过进展且不是100%进展的项目）
                one_month_ago_dt = (monday - timedelta(days=30)).strftime('%Y-%m-%d 23:59:59')
                project_sql = text("""
                    SELECT project_name, project_progress, project_status 
                    FROM project 
                    WHERE project_manager = :real_name
                      AND (project_status IS NULL OR (project_status != '已归档' AND project_status != '已结项'))
                      AND project_progress < 100.0
                      AND update_date >= :one_month_ago
                """)
                projects = conn.execute(project_sql, {
                    "real_name": real_name,
                    "one_month_ago": one_month_ago_dt
                }).mappings().all()
                
                delivery_list = []
                d_idx = 1
                if projects:
                    delivery_list.append("目前负责跟进的正在实施项目进度情况如下：")
                    for p in projects:
                        delivery_list.append(f"  {d_idx}) 项目【{p['project_name']}】当前总体进度：{float(p['project_progress'] or 0):.1f}%")
                        d_idx += 1
                
                # 7. 里程碑任务实际完成情况
                task_sql = text("""
                    SELECT t.name as task_name, p.project_name, t.milestone 
                    FROM task t
                    INNER JOIN project p ON t.project_id = p.id
                    WHERE p.project_manager = :real_name
                      AND t.finish_date BETWEEN :start_date AND :end_date
                      AND t.status = '0'
                    ORDER BY t.finish_date ASC
                """)
                tasks = conn.execute(task_sql, {
                    "real_name": real_name,
                    "start_date": start_date_str,
                    "end_date": end_date_str
                }).mappings().all()
                
                if tasks:
                    if not delivery_list:
                        delivery_list.append("本周项目子任务及里程碑节点交付动作明细：")
                    else:
                        delivery_list.append("\n本周项目子任务及里程碑节点交付动作明细：")
                    for t in tasks:
                        m_tag = "【里程碑】" if t['milestone'] == '1' else ""
                        delivery_list.append(f"  {d_idx}) {m_tag}完成项目【{t['project_name']}】下的任务节点：【{t['task_name']}】")
                        d_idx += 1
                
                formatted_delivery = perf_snapshot + ("\n".join(delivery_list) if delivery_list else "1. 本周名下负责的正在实施项目推进平稳，无重大子任务或里程碑完成提交。")
                result["delivery_actual"] = formatted_delivery
                result["sales_actual"] = ""
                
                # 8. 亮点与卡点智能提取
                highlight_list = []
                blocker_list = []
                
                h_idx = 1
                m_tasks = [t for t in tasks if t['milestone'] == '1']
                if m_tasks:
                    highlight_list.append(f"  {h_idx}) 本周成功突破并攻克了 {len(m_tasks)} 个核心项目交付里程碑节点！")
                    h_idx += 1
                
                if len(tasks) >= 3:
                    highlight_list.append(f"  {h_idx}) 本周高效推进并完成了 {len(tasks)} 个项目子项任务交付，项目稳步实施中")
                    h_idx += 1
                
                # 卡点检查
                project_block_sql = text("""
                    SELECT project_name, remarks 
                    FROM project 
                    WHERE project_manager = :real_name
                      AND stop_status = '1'
                """)
                block_projects = conn.execute(project_block_sql, {"real_name": real_name}).mappings().all()
                b_idx = 1
                if block_projects:
                    for bp in block_projects:
                        blocker_list.append(f"  {b_idx}) 交付难点：项目【{bp['project_name']}】处于暂停或异常挂起状态（备注：{bp['remarks'] or '无'}）")
                        b_idx += 1
                
                # 预设立（超过一个月未签合同）项目检查
                presetup_block_sql = text("""
                    SELECT project_name, create_date 
                    FROM project 
                    WHERE project_manager = :real_name
                      AND (project_status IS NULL OR (project_status != '已归档' AND project_status != '已结项'))
                      AND (contract_status = '0' OR contract_status IS NULL)
                      AND create_date < :one_month_ago
                """)
                presetup_projects = conn.execute(presetup_block_sql, {
                    "real_name": real_name,
                    "one_month_ago": (monday - timedelta(days=30)).strftime('%Y-%m-%d 23:59:59')
                }).mappings().all()
                if presetup_projects:
                    for pp in presetup_projects:
                        c_date_str = pp['create_date'].strftime('%Y-%m-%d') if pp['create_date'] else '—'
                        blocker_list.append(f"  {b_idx}) 预设立预警：项目【{pp['project_name']}】已立项执行超过一个月，但目前仍未签订正式合同（立项时间：{c_date_str}）")
                        b_idx += 1
                
                # 8.2 已到交付节点还未开发票的项目
                unbilled_node_sql = text("""
                    SELECT DISTINCT p.project_name, p.project_progress, np.project_progress_trigger, cm.installment_money
                    FROM project p
                    INNER JOIN contract_money_urge_notify_project np ON p.id = np.project_id
                    INNER JOIN contract_money cm ON np.contract_money_id = cm.id
                    WHERE p.project_manager = :real_name
                      AND p.project_progress >= np.project_progress_trigger
                      AND (cm.invoic_status IS NULL OR cm.invoic_status = '' OR cm.invoic_status = '0')
                      AND (p.project_status IS NULL OR (p.project_status != '已归档' AND p.project_status != '已结项'))
                """)
                unbilled_projects = conn.execute(unbilled_node_sql, {"real_name": real_name}).mappings().all()
                if unbilled_projects:
                    for up in unbilled_projects:
                        # 原始单位为“元”，需除以10000转换为“万元”
                        money_val = float(up['installment_money']) / 10000.0 if up['installment_money'] is not None else None
                        money_str = f"{money_val:,.2f}" if money_val is not None else "—"
                        blocker_list.append(
                            f"  {b_idx}) 交付卡点：项目【{up['project_name']}】进度已达 {float(up['project_progress'] or 0):.1f}%"
                            f"（已达收付款触发节点 {float(up['project_progress_trigger'] or 0):.1f}%），"
                            f"但尚未开发票（本阶段合同款项：{money_str}万元）"
                        )
                        b_idx += 1
 
                # 8.3 已开发票还未到账的项目
                unreceived_bill_sql = text("""
                    SELECT MIN(p.project_name) as project_name, br.bill_money, br.un_account_money, br.bill_create_date
                    FROM contract_un_receive_bill_not_receive br
                    INNER JOIN contract_project cp ON br.contract_id = cp.contract_id
                    INNER JOIN project p ON cp.project_id = p.id
                    WHERE p.project_manager = :real_name
                      AND br.un_account_money > 0
                      AND (p.project_status IS NULL OR (p.project_status != '已归档' AND p.project_status != '已结项'))
                    GROUP BY br.contract_id, br.bill_money, br.un_account_money, br.bill_create_date
                """)
                unreceived_projects = conn.execute(unreceived_bill_sql, {"real_name": real_name}).mappings().all()
                if unreceived_projects:
                    for urp in unreceived_projects:
                        bill_money_str = f"{float(urp['bill_money']):,.2f}" if urp['bill_money'] is not None else "—"
                        un_money_str = f"{float(urp['un_account_money']):,.2f}" if urp['un_account_money'] is not None else "—"
                        b_date_str = urp['bill_create_date'].strftime('%Y-%m-%d') if urp['bill_create_date'] else '—'
                        blocker_list.append(
                            f"  {b_idx}) 收欠款预警：项目【{urp['project_name']}】已开发票但尚未回款到账"
                            f"（开票日期：{b_date_str}，开票金额：{bill_money_str}万元，未到账金额：{un_money_str}万元）"
                        )
                        b_idx += 1
                
                result["delivery_highlights"] = "\n".join(highlight_list) if highlight_list else "1. 交付工作处于正常开发推进中，开发交付无积压。"
                result["delivery_blockers"] = warning_text + ("\n".join(blocker_list) if blocker_list else "1. 本周项目整体推进良好，暂无重大的技术难点与交付卡点。")
                result["sales_highlights"] = ""
                result["sales_blockers"] = ""

            # 无论什么岗位，都在最后将 7 个新增细粒度字段写好（如果有相应变量，就格式化；若没有或者为营销岗，默认为“—”）
            # 1. 目前负责跟进的正在实施项目进度情况
            if 'projects' in locals() and projects:
                active_list = []
                for idx, p in enumerate(projects, 1):
                    active_list.append(f"{idx}. 项目【{p['project_name']}】当前总体进度：{float(p['project_progress'] or 0):.1f}%")
                result["crm_active_projects"] = "\n".join(active_list)
            else:
                result["crm_active_projects"] = "—"
 
            # 2. 本周项目子任务及里程碑节点交付动作明细
            if 'tasks' in locals() and tasks:
                task_list = []
                for idx, t in enumerate(tasks, 1):
                    m_tag = "【里程碑】" if t['milestone'] == '1' else ""
                    task_list.append(f"{idx}. {m_tag}完成项目【{t['project_name']}】下的任务节点：【{t['task_name']}】")
                result["crm_milestone_tasks"] = "\n".join(task_list)
            else:
                result["crm_milestone_tasks"] = "—"
 
            # 3. 处于暂停或异常挂起状态的项目
            if 'block_projects' in locals() and block_projects:
                suspended_list = []
                for idx, bp in enumerate(block_projects, 1):
                    suspended_list.append(f"{idx}. 项目【{bp['project_name']}】处于暂停或异常挂起状态（备注：{bp['remarks'] or '无'}）")
                result["crm_suspended_projects"] = "\n".join(suspended_list)
            else:
                result["crm_suspended_projects"] = "—"
 
            # 4. 项目已立项执行超过一个月，但目前仍未签订正式合同
            if 'presetup_projects' in locals() and presetup_projects:
                no_contract_list = []
                for idx, pp in enumerate(presetup_projects, 1):
                    c_date_str = pp['create_date'].strftime('%Y-%m-%d') if pp['create_date'] else '—'
                    no_contract_list.append(f"{idx}. 项目【{pp['project_name']}】已立项执行超过一个月，但目前仍未签订正式合同（立项时间：{c_date_str}）")
                result["crm_no_contract_warning"] = "\n".join(no_contract_list)
            else:
                result["crm_no_contract_warning"] = "—"
 
            # 5. 交付卡点：项目有进度但尚未开发票
            if 'unbilled_projects' in locals() and unbilled_projects:
                unbilled_list = []
                for idx, up in enumerate(unbilled_projects, 1):
                    money_val = float(up['installment_money']) / 10000.0 if up['installment_money'] is not None else None
                    money_str = f"{money_val:,.2f}" if money_val is not None else "—"
                    unbilled_list.append(
                        f"{idx}. 项目【{up['project_name']}】进度已达 {float(up['project_progress'] or 0):.1f}%"
                        f"（已达收付款触发节点 {float(up['project_progress_trigger'] or 0):.1f}%），"
                        f"但尚未开发票（本阶段合同款项：{money_str}万元）"
                    )
                result["crm_unbilled_warning"] = "\n".join(unbilled_list)
            else:
                result["crm_unbilled_warning"] = "—"
 
            # 6. 收欠款预警：项目已开发票但尚未回款到账
            if 'unreceived_projects' in locals() and unreceived_projects:
                unreceived_list = []
                for idx, urp in enumerate(unreceived_projects, 1):
                    bill_money_str = f"{float(urp['bill_money']):,.2f}" if urp['bill_money'] is not None else "—"
                    un_money_str = f"{float(urp['un_account_money']):,.2f}" if urp['un_account_money'] is not None else "—"
                    b_date_str = urp['bill_create_date'].strftime('%Y-%m-%d') if urp['bill_create_date'] else '—'
                    unreceived_list.append(
                        f"{idx}. 项目【{urp['project_name']}】已开发票但尚未回款到账"
                        f"（开票日期：{b_date_str}，开票金额：{bill_money_str}万元，未到账金额：{un_money_str}万元）"
                    )
                result["crm_unreceived_warning"] = "\n".join(unreceived_list)
            else:
                result["crm_unreceived_warning"] = "—"
 
            # 7. 个人工作饱和度与项目健康度诊断
            if 'personal_warnings' in locals() and personal_warnings:
                result["crm_health_diagnosis"] = "\n".join([f"{idx}. {w}" for idx, w in enumerate(personal_warnings, 1)])
            else:
                result["crm_health_diagnosis"] = "1. 工作饱和度与项目实施状态正常，暂无诊断预警。"
                
    except Exception as e:
        import logging
        logging.getLogger("battle100").error(f"从 CRM 库提取数据出错: {e}")
        
    return result


@router.get("/weekly/auto-extract-crm", summary="自动从CRM系统提取周报实际完成和达成情况")
async def extract_weekly_crm_data(
    start_date: date = Query(..., description="周开始日期(周一)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    联动 CRM 数据库，基于当前登录用户的真实姓名（real_name）自动抽取其本周的：
    1. 新签合同与回款明细（营销岗）
    2. 客户拜访跟进频次（营销岗）
    3. 月度新签与回款目标的最新达成率说明（自动计算）
    4. 正在进行的正在实施项目进度及本周攻克的里程碑子任务（交付岗）
    并自动拼装排版为序号（1. 2. 3.）明细文本，提供一键拉取静默覆盖。
    """
    from fastapi.concurrency import run_in_threadpool
    from app.models.user import PositionType, UserRole
    
    # 岗位判定
    is_marketing = (
        current_user.position_type == PositionType.MARKETING 
        or current_user.role in [UserRole.TARGET_OFFICER, UserRole.MARKETING_STAFF, UserRole.TECH_MARKETING]
    )
    
    # 异步在线程池中调用同步数据库查询
    result = await run_in_threadpool(
        sync_extract_crm_data, 
        current_user.name, # 绑定真实姓名进行跨库匹配
        start_date, 
        is_marketing
    )
    
    return result


@router.get("/weekly/summary", response_model=WeeklyReportListResponse, summary="获取小组成员周复盘汇总表")
async def get_weekly_reports_summary(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """获取选定周时间段内，小组成员提交的周复盘数据大表（PC端汇总使用）"""
    from app.models.user import UserRole
    if current_user.role not in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value] and current_user.team_id is not None:
        team_id = current_user.team_id

    from app.models.user import User as DbUser
    
    query = select(
        WeeklyReport.id,
        WeeklyReport.user_id,
        WeeklyReport.start_date,
        WeeklyReport.end_date,
        WeeklyReport.delivery_plan,
        WeeklyReport.sales_plan,
        WeeklyReport.delivery_actual,
        WeeklyReport.sales_actual,
        WeeklyReport.delivery_rate,
        WeeklyReport.sales_rate,
        WeeklyReport.delivery_highlights,
        WeeklyReport.sales_highlights,
        WeeklyReport.delivery_blockers,
        WeeklyReport.sales_blockers,
        WeeklyReport.delivery_support,
        WeeklyReport.sales_support,
        WeeklyReport.next_delivery_plan,
        WeeklyReport.next_sales_plan,
        WeeklyReport.status,
        WeeklyReport.submitted_at,
        WeeklyReport.created_at,
        WeeklyReport.updated_at,
        DbUser.name.label("user_name"),
        DbUser.position_type.label("user_position_type"),
        DbUser.role.label("user_role")
    ).join(DbUser, WeeklyReport.user_id == DbUser.id)
    
    if team_id:
        query = query.where(DbUser.team_id == team_id)
        
    if third_class_bar and third_class_bar != "all":
        query = query.where(DbUser.third_class_bar == third_class_bar)
        
    if user_name:
        if "," in user_name or "，" in user_name:
            from sqlalchemy import or_
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                query = query.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            query = query.where(DbUser.name.like(f"%{user_name}%"))

        
    query = query.where(WeeklyReport.start_date == start_date)
    query = query.order_by(DbUser.name.asc())
    
    # 统计总数
    count_query = select(func.count()).select_from(query.subquery())
    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0
    
    # 分页
    query = query.offset((page - 1) * page_size).limit(page_size)
    res = await db.execute(query)
    rows = res.all()
    
    items = []
    for row in rows:
        items.append(WeeklyReportResponse(
            id=row.id,
            user_id=row.user_id,
            start_date=row.start_date,
            end_date=row.end_date,
            delivery_plan=row.delivery_plan,
            sales_plan=row.sales_plan,
            delivery_actual=row.delivery_actual,
            sales_actual=row.sales_actual,
            delivery_rate=row.delivery_rate,
            sales_rate=row.sales_rate,
            delivery_highlights=row.delivery_highlights,
            sales_highlights=row.sales_highlights,
            delivery_blockers=row.delivery_blockers,
            sales_blockers=row.sales_blockers,
            delivery_support=row.delivery_support,
            sales_support=row.sales_support,
            next_delivery_plan=row.next_delivery_plan,
            next_sales_plan=row.next_sales_plan,
            status=row.status,
            submitted_at=row.submitted_at,
            created_at=row.created_at,
            updated_at=row.updated_at,
            user_name=row.user_name,
            user_position_type=row.user_position_type,
            user_role=row.user_role
        ))
        
    return WeeklyReportListResponse(total=total, items=items)


@router.get("/weekly/crm-summary", response_model=WeeklyCrmSummaryListResponse, summary="获取全员/小组成员 CRM 业务数据汇总")
async def get_weekly_reports_crm_summary(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """
    联动 CRM 系统，根据当前登录用户的角色，智能并发拉取可见范围内所有成员的当周 CRM 业绩与进度。
    1. 系统管理员和目标官可看全员（支持按战队筛选）。
    2. 战队长可看同三级巴的队员，以及同战队成员。
    3. 队员可看同三级巴或同战队成员。
    """
    from sqlalchemy import or_, and_
    from app.models.user import User as DbUser, PositionType
    from app.models.organization import Team as DbTeam
    from fastapi.concurrency import run_in_threadpool
    import asyncio

    # 1. 确定当前用户可看的人员范围
    stmt = select(
        DbUser.id,
        DbUser.name,
        DbUser.third_class_bar,
        DbUser.position_type,
        DbUser.role,
        DbTeam.name.label("team_name")
    ).outerjoin(DbTeam, DbUser.team_id == DbTeam.id).where(DbUser.is_active == True)

    count_stmt = select(func.count(DbUser.id)).where(DbUser.is_active == True)

    if current_user.role in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value]:
        # 管理员和目标官：可看所有人。支持前端筛选 team_id
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
            count_stmt = count_stmt.where(DbUser.team_id == team_id)
    else:
        # 非全局角色（包括战队长、队员）：只能看同战队或同三级巴成员，以及当前用户自己
        conditions = [DbUser.id == current_user.id]
        if current_user.team_id is not None:
            conditions.append(DbUser.team_id == current_user.team_id)
        if current_user.third_class_bar:
            conditions.append(and_(
                DbUser.third_class_bar == current_user.third_class_bar,
                DbUser.third_class_bar != None,
                DbUser.third_class_bar != ""
            ))
        stmt = stmt.where(or_(*conditions))
        count_stmt = count_stmt.where(or_(*conditions))
        
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
            count_stmt = count_stmt.where(DbUser.team_id == team_id)

    if third_class_bar and third_class_bar != "all":
        stmt = stmt.where(DbUser.third_class_bar == third_class_bar)
        count_stmt = count_stmt.where(DbUser.third_class_bar == third_class_bar)

    if user_name:
        if "," in user_name or "，" in user_name:
            from sqlalchemy import or_
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                stmt = stmt.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
                count_stmt = count_stmt.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            stmt = stmt.where(DbUser.name.like(f"%{user_name}%"))
            count_stmt = count_stmt.where(DbUser.name.like(f"%{user_name}%"))


    # 2. 统计总数
    total_res = await db.execute(count_stmt)
    total = total_res.scalar() or 0

    # 3. 分页并按姓名排序
    stmt = stmt.order_by(DbUser.name.asc()).offset((page - 1) * page_size).limit(page_size)
    res = await db.execute(stmt)
    rows = res.all()

    # 4. 线程池并发执行各个用户的 CRM 数据拉取
    async def fetch_user_crm_data(user_row):
        is_marketing = (
            user_row.position_type == PositionType.MARKETING 
            or user_row.role in [UserRole.TARGET_OFFICER, UserRole.MARKETING_STAFF, UserRole.TECH_MARKETING]
        )
        # 调用跨库查询函数
        crm_data = await run_in_threadpool(
            sync_extract_crm_data, 
            user_row.name,
            start_date, 
            is_marketing
        )
        return WeeklyCrmSummaryItem(
            user_id=user_row.id,
            user_name=user_row.name,
            third_class_bar=user_row.third_class_bar,
            team_name=user_row.team_name,
            position_type=user_row.position_type.value if user_row.position_type else None,
            role=user_row.role.value if hasattr(user_row.role, "value") else str(user_row.role),
            delivery_actual=crm_data.get("delivery_actual"),
            sales_actual=crm_data.get("sales_actual"),
            delivery_rate=crm_data.get("delivery_rate"),
            sales_rate=crm_data.get("sales_rate"),
            delivery_highlights=crm_data.get("delivery_highlights"),
            sales_highlights=crm_data.get("sales_highlights"),
            delivery_blockers=crm_data.get("delivery_blockers"),
            sales_blockers=crm_data.get("sales_blockers"),
            crm_active_projects=crm_data.get("crm_active_projects"),
            crm_milestone_tasks=crm_data.get("crm_milestone_tasks"),
            crm_suspended_projects=crm_data.get("crm_suspended_projects"),
            crm_no_contract_warning=crm_data.get("crm_no_contract_warning"),
            crm_unbilled_warning=crm_data.get("crm_unbilled_warning"),
            crm_unreceived_warning=crm_data.get("crm_unreceived_warning"),
            crm_health_diagnosis=crm_data.get("crm_health_diagnosis")
        )

    tasks = [fetch_user_crm_data(row) for row in rows]
    items = await asyncio.gather(*tasks) if tasks else []

    return WeeklyCrmSummaryListResponse(total=total, items=items)


@router.delete("/weekly/{report_id}", summary="删除指定周报")
async def delete_weekly_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """物理删除某条周报数据，需权限校验"""
    stmt = select(WeeklyReport).where(WeeklyReport.id == report_id)
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="未找到该周报记录")
        
    # 权限校验：仅管理员、拥有 delete_weekly_report 权限的角色，或周报所有者本人允许删除
    from app.models.user import UserRole, RolePermission
    has_perm = current_user.role == UserRole.ADMIN.value
    if not has_perm:
        perm_res = await db.execute(
            select(RolePermission).where(
                RolePermission.role == current_user.role,
                RolePermission.menu_key == "delete_weekly_report"
            )
        )
        if perm_res.scalar_one_or_none():
            has_perm = True

    if not has_perm and report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除他人的周报记录")
        
    await db.delete(report)
    await db.commit()
    return {"message": "周报已成功删除"}


from pydantic import BaseModel

class WeeklyBatchDeleteRequest(BaseModel):
    """批量删除请求Schema"""
    ids: list[int]


@router.post("/weekly/batch-delete", summary="批量删除周报记录")
async def batch_delete_weekly_reports(
    req: WeeklyBatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """批量物理删除多条周报数据，仅管理员与拥有 delete_weekly_report 权限的角色允许操作"""
    from app.models.user import UserRole, RolePermission
    has_perm = current_user.role == UserRole.ADMIN.value
    if not has_perm:
        perm_res = await db.execute(
            select(RolePermission).where(
                RolePermission.role == current_user.role,
                RolePermission.menu_key == "delete_weekly_report"
            )
        )
        if perm_res.scalar_one_or_none():
            has_perm = True

    if not has_perm:
        raise HTTPException(status_code=403, detail="仅管理员或拥有删除周报权限的角色可以批量删除周报")
        
    from sqlalchemy import delete
    stmt = delete(WeeklyReport).where(WeeklyReport.id.in_(req.ids))
    await db.execute(stmt)
    await db.commit()
    return {"message": f"成功删除 {len(req.ids)} 条周报记录"}


@router.put("/weekly/{report_id}", response_model=WeeklyReportResponse, summary="修改/更新指定周报")
async def update_weekly_report(
    report_id: int,
    report_in: WeeklyReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """修改/更新已有周报内容（支持管理员/拥有 edit_weekly_report 权限者修改他人，或所有人修改自己）"""
    stmt = select(WeeklyReport).where(WeeklyReport.id == report_id)
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="未找到该周报记录")
        
    # 权限校验：仅管理员、拥有 edit_weekly_report 权限的角色，或周报所有者本人允许修改
    from app.models.user import UserRole, RolePermission
    has_perm = current_user.role == UserRole.ADMIN.value
    if not has_perm:
        perm_res = await db.execute(
            select(RolePermission).where(
                RolePermission.role == current_user.role,
                RolePermission.menu_key == "edit_weekly_report"
            )
        )
        if perm_res.scalar_one_or_none():
            has_perm = True

    if not has_perm and report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权修改他人的周报记录")
        
    update_data = report_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(report, field, value)
        
    db.add(report)
    await db.commit()
    await db.refresh(report)
    
    # 关联补全被编辑者的用户信息
    from app.models.user import User as DbUser
    user_stmt = select(DbUser).where(DbUser.id == report.user_id)
    user_res = await db.execute(user_stmt)
    db_user = user_res.scalar_one()
    
    report.user_name = db_user.name
    report.user_position_type = db_user.position_type
    report.user_role = db_user.role
    
    return report


# ==========================================
#          战队/三级巴 整体周报与产值统计 API
# ==========================================

def sync_get_group_crm_data(user_names: list[str], crm_user_ids: list[str], start_date_val: date, is_company_wide: bool = False, group_name: str | None = None) -> dict:
    """同步直连 CRM 数据库统计指定成员周内的产值、线索、回款及项目反馈明细"""
    from datetime import timedelta
    from sqlalchemy import text
    from app.database import get_crm_db
    
    start_date_str = start_date_val.strftime('%Y-%m-%d 00:00:00')
    end_date_str = (start_date_val + timedelta(days=6)).strftime('%Y-%m-%d 23:59:59')
    
    res = {
        "potential_leads": 0,
        "valid_leads": 0,
        "production_value": 0.0,
        "receive_value": 0.0,
        "crm_details_text": "暂无 CRM 沟通或项目跟进明细"
    }
    
    try:
        with get_crm_db() as conn:
            # 1. 统计 CRM 有效线索 (25%) 与潜力线索 (5%-10%)
            if crm_user_ids:
                user_ids_str = ", ".join([f"'{uid}'" for uid in crm_user_ids])
                
                # 潜力商机线索
                pot_sql = text(f"""
                    SELECT COUNT(*) as count 
                    FROM zdcrm_business_opportunity 
                    WHERE progress BETWEEN 5 AND 10 
                      AND (is_suspension = '0' OR is_suspension IS NULL)
                      AND market_user_id IN ({user_ids_str})
                      AND update_time BETWEEN :start AND :end
                """)
                pot_val = conn.execute(pot_sql, {"start": start_date_str, "end": end_date_str}).scalar() or 0
                res["potential_leads"] = int(pot_val)
                
                # 有效商机线索
                val_sql = text(f"""
                    SELECT COUNT(*) as count 
                    FROM zdcrm_business_opportunity 
                    WHERE progress = 25 
                      AND (is_suspension = '0' OR is_suspension IS NULL)
                      AND market_user_id IN ({user_ids_str})
                      AND update_time BETWEEN :start AND :end
                """)
                val_val = conn.execute(val_sql, {"start": start_date_str, "end": end_date_str}).scalar() or 0
                res["valid_leads"] = int(val_val)
                
            # 2. 统计 CRM 累计产值额（万元，过滤 createDate 落在当周，负责人为本组成员。同时排除历史旧账期跑批数据的虚高干扰）
            if user_names:
                from datetime import timedelta
                names_str = ", ".join([f"'{name}'" for name in user_names])
                month_start_date = start_date_val.replace(day=1)
                prev_month_start_date = (month_start_date - timedelta(days=15)).replace(day=1)
                
                prod_sql = text(f"""
                    SELECT COALESCE(SUM(dp.money), 0) as total_prod
                    FROM dashboard_production dp
                    JOIN project p ON dp.project_id = p.id
                    WHERE p.project_manager IN ({names_str})
                      AND dp.createDate BETWEEN :start AND :end
                      AND dp.account_date IN (:prev_month_start, :month_start)
                      AND dp.isDel = '0'
                """)
                prod_val = conn.execute(prod_sql, {
                    "start": start_date_str,
                    "end": end_date_str,
                    "prev_month_start": prev_month_start_date.strftime('%Y-%m-%d'),
                    "month_start": month_start_date.strftime('%Y-%m-%d')
                }).scalar() or 0.0
                res["production_value"] = float(prod_val) / 10000.0 # 元转万元
                
                # 3. 统计 CRM 实际到账回款额（万元，过滤回款落在当周）
                if is_company_wide:
                    # 全公司大盘：去除人员归属限制，拉取全量到账流水，所有注释必须使用中文
                    recv_sql = text(f"""
                        SELECT COALESCE(SUM(r.receive_money), 0) as total_recv
                        FROM zdcrm_contract_receive_money_view r
                        INNER JOIN contract c ON r.contract_id = c.id
                        WHERE r.create_date BETWEEN :start_date AND :end_date
                    """)
                else:
                    # 战队/小组级视角：改为支持合同所属部门进行过滤。若匹配不到战队名称则fallback至原有人员联合过滤
                    # 所有注释必须使用中文
                    dept_conditions = None
                    if group_name == "广州三战队（大数据）":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD007,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD007,%%' OR c.office_code = 'ZD007')"
                    elif group_name == "清远战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD020,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD020,%%' OR c.office_code = 'ZD020')"
                    elif group_name == "广州二战队":
                        dept_conditions = "((o1.parent_codes LIKE '%%,ZD004,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD004,%%' OR c.office_code = 'ZD004') AND c.office_code NOT IN ('bf6857845da54c72bb1df29cf6289fda', '402881e48695f0e201869ae61ddf05bd', '402881e4860c74ea01862edc56191199'))"
                    elif group_name == "广州一战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD019,%%' OR o1.parent_codes LIKE '%%,ZD006,%%' OR o1.parent_codes LIKE '%%,ZD003,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD019,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD006,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD003,%%' OR c.office_code IN ('ZD019', 'ZD006', 'ZD003') OR o2.ORGANIZTREECODES LIKE '%%,bfaae30eec904c879d15bf10a3c04aa2,%%')"
                    elif group_name == "佛山战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD013,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD013,%%' OR c.office_code = 'ZD013')"
                    elif group_name == "东莞战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,402881e47d4b00d1017d4b97bca2000b,%%' OR o2.ORGANIZTREECODES LIKE '%%,402881e47d4b00d1017d4b97bca2000b,%%' OR c.office_code = '402881e47d4b00d1017d4b97bca2000b' OR c.office_code = 'bf6857845da54c72bb1df29cf6289fda' OR o2.ORGANIZTREECODES LIKE '%%,bf6857845da54c72bb1df29cf6289fda,%%')"
                    elif group_name == "湛江战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD012,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD012,%%' OR c.office_code = 'ZD012' OR c.office_code = '402881e4860c74ea01862edc56191199' OR o2.ORGANIZTREECODES LIKE '%%,402881e4860c74ea01862edc56191199,%%')"
                    elif group_name == "云浮战队":
                        dept_conditions = "(o1.parent_codes LIKE '%%,ZD011,%%' OR o1.parent_codes LIKE '%%,ZD010,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD011,%%' OR o2.ORGANIZTREECODES LIKE '%%,ZD010,%%' OR c.office_code IN ('ZD011', 'ZD010', '402881e48695f0e201869ae61ddf05bd') OR o2.ORGANIZTREECODES LIKE '%%,402881e48695f0e201869ae61ddf05bd,%%')"
                    elif group_name == "茂名战队":
                        dept_conditions = "(c.office_code IN ('402881e48185cf30018185d67353000c', '402881e4838933c20183c4fc1e0413e8', 'dd8e09fdf59e48e7af220b917436b71e') OR o2.ORGANIZTREECODES LIKE '%%,dd8e09fdf59e48e7af220b917436b71e,%%')"
                    
                    if dept_conditions:
                        recv_sql = text(f"""
                            SELECT COALESCE(SUM(r.receive_money), 0) as total_recv
                            FROM zdcrm_contract_receive_money_view r
                            INNER JOIN contract c ON r.contract_id = c.id
                            LEFT JOIN js_sys_office o1 ON c.office_code = o1.office_code
                            LEFT JOIN (
                                SELECT o1.ORGANIZCODE, o1.ORGANIZTREECODES, o1.ORGANIZNAME
                                FROM zdcrm_history_organization o1
                                INNER JOIN (
                                    SELECT ORGANIZCODE, MAX(year) as max_year
                                    FROM zdcrm_history_organization
                                    GROUP BY ORGANIZCODE
                                ) o2 ON o1.ORGANIZCODE = o2.ORGANIZCODE AND o1.year = o2.max_year
                            ) o2 ON c.office_code = o2.ORGANIZCODE
                            WHERE ({dept_conditions})
                              AND r.create_date BETWEEN :start_date AND :end_date
                        """)
                    else:
                        # Fallback 到原有的按人员联合过滤
                        conditions = []
                        if user_names:
                            names_str_escaped = ", ".join([f"'{name}'" for name in user_names])
                            conditions.append(f"c.signer IN ({names_str_escaped})")
                            conditions.append(f"c.contract_head_user IN ({names_str_escaped})")
                        if crm_user_ids:
                            crm_user_ids_str_escaped = ", ".join([f"'{uid}'" for uid in crm_user_ids if uid])
                            conditions.append(f"c.create_by IN ({crm_user_ids_str_escaped})")
                        
                        if conditions:
                            conditions_sql = " OR ".join(conditions)
                            recv_sql = text(f"""
                                SELECT COALESCE(SUM(r.receive_money), 0) as total_recv
                                FROM zdcrm_contract_receive_money_view r
                                INNER JOIN contract c ON r.contract_id = c.id
                                WHERE ({conditions_sql})
                                  AND r.create_date BETWEEN :start_date AND :end_date
                            """)
                        else:
                            recv_sql = None
                
                if recv_sql:
                    recv_val = conn.execute(recv_sql, {
                        "start_date": start_date_str,
                        "end_date": end_date_str
                    }).scalar() or 0.0
                    res["receive_value"] = float(recv_val)
                
                # 4. 放宽 Token 至 20K-30K 后，拉取客户拜访与进度反馈原文等真实细节，喂给 AI 大模型
                detail_texts = []
                
                # 本周商务跟进拜访记录原文
                visit_sql = text(f"""
                    SELECT customer_name, remark, create_time, create_by
                    FROM zdcrm_visit_customer_record
                    WHERE create_by IN ({names_str})
                      AND create_time BETWEEN :start AND :end
                      AND is_del = '0'
                    ORDER BY create_time ASC
                    LIMIT 20
                """)
                visits = conn.execute(visit_sql, {"start": start_date_str, "end": end_date_str}).mappings().all()
                if visits:
                    detail_texts.append("【成员商务拜访/客户对接详情记录】:")
                    for idx, v in enumerate(visits, 1):
                        t_str = v['create_time'].strftime('%m-%d') if v['create_time'] else '—'
                        detail_texts.append(f"  {idx}) 成员[{v['create_by']}]于 {t_str} 对接客户【{v['customer_name']}】，跟进记录：“{v['remark'] or '未填'}”")
                
                # 本周进度变动项目及其确认产值详情
                p_change_sql = text(f"""
                    SELECT DISTINCT p.project_name, p.project_manager, dp.start_progress, dp.end_progress, dp.progress_change, dp.money
                    FROM dashboard_production dp
                    JOIN project p ON dp.project_id = p.id
                    WHERE p.project_manager IN ({names_str})
                      AND dp.createDate BETWEEN :start AND :end
                      AND dp.account_date IN (:prev_month_start, :month_start)
                      AND dp.isDel = '0'
                    LIMIT 20
                """)
                p_changes = conn.execute(p_change_sql, {
                    "start": start_date_str,
                    "end": end_date_str,
                    "prev_month_start": prev_month_start_date.strftime('%Y-%m-%d'),
                    "month_start": month_start_date.strftime('%Y-%m-%d')
                }).mappings().all()
                if p_changes:
                    detail_texts.append("\n【项目交付进度异动与确认产值记录】:")
                    for idx, pc in enumerate(p_changes, 1):
                        detail_texts.append(
                            f"  {idx}) 负责人[{pc['project_manager']}]负责的【{pc['project_name']}】"
                            f"进度由 {pc['start_progress']}% 推进至 {pc['end_progress']}%"
                            f"（异动：+{pc['progress_change']}%，确认产值额 {float(pc['money'])/10000:.2f} 万元）"
                        )
                
                # 交付难点 (已到交付节点但未开票)
                unbilled_sql = text(f"""
                    SELECT DISTINCT p.project_name, p.project_manager, p.project_progress, np.project_progress_trigger, cm.installment_money
                    FROM project p
                    INNER JOIN contract_money_urge_notify_project np ON p.id = np.project_id
                    INNER JOIN contract_money cm ON np.contract_money_id = cm.id
                    WHERE p.project_manager IN ({names_str})
                      AND p.project_progress >= np.project_progress_trigger
                      AND (cm.invoic_status IS NULL OR cm.invoic_status = '' OR cm.invoic_status = '0')
                      AND (p.project_status IS NULL OR (p.project_status != '已归档' AND p.project_status != '已结项'))
                    LIMIT 20
                """)
                unbilled = conn.execute(unbilled_sql).mappings().all()
                if unbilled:
                    detail_texts.append("\n【交付卡点预警（已达交付触发节点但尚未开票）】:")
                    for idx, ub in enumerate(unbilled, 1):
                        # 原始单位为“元”，需除以10000转换为“万元”
                        money_val = float(ub['installment_money']) / 10000.0 if ub['installment_money'] is not None else 0.0
                        detail_texts.append(
                            f"  {idx}) 项目【{ub['project_name']}】(负责人: {ub['project_manager']}) "
                            f"当前进度已达 {float(ub['project_progress']):.1f}% (已触发节点 {float(ub['project_progress_trigger']):.1f}%)，"
                            f"但未开具发票，影响后续收款回款。对应阶段款项金额: {money_val:.2f} 万元。"
                        )
                
                # 本周活跃正在实施项目清单 (用于 AI 诊断成员工作饱和度与项目空仓/断粮风险)
                active_projects_sql = text(f"""
                    SELECT project_name, project_manager, project_progress, project_status
                    FROM project
                    WHERE project_manager IN ({names_str})
                      AND project_progress < 100.0
                      AND (project_status IS NULL OR (project_status != '已归档' AND project_status != '已结项' AND project_status != '3'))
                    LIMIT 40
                """)
                active_projects = conn.execute(active_projects_sql).mappings().all()
                if active_projects:
                    detail_texts.append("\n【全组各成员名下负责的所有活跃正在实施项目及当前最新进度清单】:")
                    for idx, ap in enumerate(active_projects, 1):
                        progress_val = ap['project_progress']
                        try:
                            progress_val_float = float(progress_val) if progress_val is not None else 0.0
                            progress_str = f"{progress_val_float:.1f}%"
                        except Exception:
                            progress_str = f"{progress_val or 0}%"
                        
                        detail_texts.append(
                            f"  {idx}) 负责人[{ap['project_manager']}]负责的活跃项目【{ap['project_name']}】"
                            f"(当前进度：{progress_str}，状态：{ap['project_status'] or '进行中'})"
                        )
                
                if detail_texts:
                    res["crm_details_text"] = "\n".join(detail_texts)
                    
    except Exception as e:
        import logging
        logging.getLogger("battle100").warning(f"同步拉取团队 CRM 数据细节异常: {e}")
        
    return res


async def get_group_weekly_metrics(
    db: AsyncSession, 
    start_date: date, 
    team_id: int | None, 
    third_class_bar: str | None
) -> dict:
    """提取战队或三级巴的本地多维业务动作，并并发直连获取 CRM 产值回款数据"""
    from datetime import timedelta, datetime, time, timezone
    from app.models.user import User as DbUser, PositionType
    from app.models.organization import Team as DbTeam
    from app.models.broadcast import BroadcastEvent
    from app.models.report import DetailType
    from sqlalchemy.orm import aliased
    from fastapi.concurrency import run_in_threadpool
    
    end_date = start_date + timedelta(days=6)
    
    start_datetime = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
    end_datetime = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
    
    # 1. 查找小组名称及全员激活成员
    group_name = "团队"
    if team_id:
        t_res = await db.execute(select(DbTeam.name).where(DbTeam.id == team_id))
        group_name = t_res.scalar() or "指定战队"
    elif third_class_bar and third_class_bar != "all":
        group_name = f"{third_class_bar}三级巴"
        
    # 查询本组所有激活成员
    user_stmt = select(DbUser.id, DbUser.name, DbUser.crm_user_id).where(DbUser.is_active == True)
    if team_id:
        user_stmt = user_stmt.where(DbUser.team_id == team_id)
    if third_class_bar and third_class_bar != "all":
        user_stmt = user_stmt.where(DbUser.third_class_bar == third_class_bar)
        
    res = await db.execute(user_stmt)
    members = res.all()
    
    user_ids = [m.id for m in members]
    user_names = [m.name for m in members]
    crm_user_ids = [m.crm_user_id for m in members if m.crm_user_id]
    
    metrics = {
        "marketing_signed": 0.0,
        "delivery_signed": 0.0,
        "win_bids": 0,
        "happiness_count": 0,
        "triangle_count": 0,
        "valid_leads": 0,
        "potential_leads": 0,
        "production_value": 0.0,
        "receive_value": 0.0
    }
    
    if not user_ids:
        return {
            "group_name": group_name,
            "user_ids": [],
            "user_names": [],
            "members_list": [],
            "metrics": metrics,
            "crm_details_text": "暂无成员数据"
        }
        
    # 2. 本系统业务指标汇总
    PartnerUser = aliased(DbUser)
    
    # 营销新签合同额
    m_signed_stmt = select(func.coalesce(func.sum(ReportDetail.amount), 0)).select_from(ReportDetail)\
        .join(DailyReport, ReportDetail.report_id == DailyReport.id)\
        .join(DbUser, DailyReport.user_id == DbUser.id)\
        .outerjoin(PartnerUser, ReportDetail.partner_user_id == PartnerUser.id)\
        .where(
            DailyReport.status == ReportStatus.REVIEWED,
            DailyReport.report_date >= start_date,
            DailyReport.report_date <= end_date,
            ReportDetail.detail_type == DetailType.CONTRACT,
            (
                ((ReportDetail.description.contains("营销新签分摊")) & ((DbUser.id.in_(user_ids)) | (PartnerUser.id.in_(user_ids)))) |
                ((~ReportDetail.description.contains("交付新签分摊")) & (
                    ((DbUser.id.in_(user_ids)) & (DbUser.position_type.in_([PositionType.MARKETING, PositionType.MANAGEMENT]))) |
                    ((PartnerUser.id.in_(user_ids)) & (PartnerUser.position_type.in_([PositionType.MARKETING, PositionType.MANAGEMENT])))
                ))
            )
        )
    m_signed_res = await db.execute(m_signed_stmt)
    metrics["marketing_signed"] = float(m_signed_res.scalar() or 0.0)
    
    # 交付新签合同额
    d_signed_stmt = select(func.coalesce(func.sum(ReportDetail.amount), 0)).select_from(ReportDetail)\
        .join(DailyReport, ReportDetail.report_id == DailyReport.id)\
        .join(DbUser, DailyReport.user_id == DbUser.id)\
        .outerjoin(PartnerUser, ReportDetail.partner_user_id == PartnerUser.id)\
        .where(
            DailyReport.status == ReportStatus.REVIEWED,
            DailyReport.report_date >= start_date,
            DailyReport.report_date <= end_date,
            ReportDetail.detail_type == DetailType.CONTRACT,
            (
                ((ReportDetail.description.contains("交付新签分摊")) & ((DbUser.id.in_(user_ids)) | (PartnerUser.id.in_(user_ids)))) |
                ((~ReportDetail.description.contains("营销新签分摊")) & (
                    ((DbUser.id.in_(user_ids)) & (DbUser.position_type.in_([PositionType.TECHNICAL, PositionType.DELIVERY]))) |
                    ((PartnerUser.id.in_(user_ids)) & (PartnerUser.position_type.in_([PositionType.TECHNICAL, PositionType.DELIVERY])))
                ))
            )
        )
    d_signed_res = await db.execute(d_signed_stmt)
    metrics["delivery_signed"] = float(d_signed_res.scalar() or 0.0)
    
    # 中标个数
    win_bids_stmt = select(func.count(BroadcastEvent.id)).where(
        BroadcastEvent.user_id.in_(user_ids),
        BroadcastEvent.event_type == "lead_75",
        BroadcastEvent.event_time >= start_datetime,
        BroadcastEvent.event_time <= end_datetime,
        BroadcastEvent.is_deleted == False
    )
    win_bids_res = await db.execute(win_bids_stmt)
    metrics["win_bids"] = int(win_bids_res.scalar() or 0)
    
    # 幸福行动数
    happiness_stmt = select(func.count(ReportDetail.id)).join(DailyReport, ReportDetail.report_id == DailyReport.id).where(
        DailyReport.status == ReportStatus.REVIEWED,
        DailyReport.report_date >= start_date,
        DailyReport.report_date <= end_date,
        ReportDetail.detail_type == DetailType.HAPPINESS,
        DailyReport.user_id.in_(user_ids)
    )
    happiness_res = await db.execute(happiness_stmt)
    metrics["happiness_count"] = int(happiness_res.scalar() or 0)
    
    # 铁三角联动次数
    triangle_stmt = select(func.count(ReportDetail.id)).join(DailyReport, ReportDetail.report_id == DailyReport.id).where(
        DailyReport.status == ReportStatus.REVIEWED,
        DailyReport.report_date >= start_date,
        DailyReport.report_date <= end_date,
        ReportDetail.detail_type == DetailType.TRIANGLE,
        DailyReport.user_id.in_(user_ids)
    )
    triangle_res = await db.execute(triangle_stmt)
    metrics["triangle_count"] = int(triangle_res.scalar() or 0)
    
    # 本地有效线索
    leads_stmt = select(func.count(ReportDetail.id)).join(DailyReport, ReportDetail.report_id == DailyReport.id).where(
        DailyReport.status == ReportStatus.REVIEWED,
        DailyReport.report_date >= start_date,
        DailyReport.report_date <= end_date,
        ReportDetail.detail_type == DetailType.LEAD,
        DailyReport.user_id.in_(user_ids),
        (ReportDetail.lead_progress.contains("25") | (ReportDetail.lead_progress == "25%"))
    )
    leads_res = await db.execute(leads_stmt)
    local_valid = int(leads_res.scalar() or 0)

    # 本地潜在线索确定
    potential_leads_stmt = select(func.count(ReportDetail.id)).join(DailyReport, ReportDetail.report_id == DailyReport.id).where(
        DailyReport.status == ReportStatus.REVIEWED,
        DailyReport.report_date >= start_date,
        DailyReport.report_date <= end_date,
        ReportDetail.detail_type == DetailType.POTENTIAL_LEAD,
        DailyReport.user_id.in_(user_ids)
    )
    potential_leads_res = await db.execute(potential_leads_stmt)
    local_potential = int(potential_leads_res.scalar() or 0)
    
    # 3. CRM 并发拉取
    is_company_wide = (team_id is None) and (third_class_bar is None or third_class_bar == "all")
    crm_res = await run_in_threadpool(
        sync_get_group_crm_data,
        user_names,
        crm_user_ids,
        start_date,
        is_company_wide,
        group_name
    )
    
    # 合并有效线索与潜力线索：本地录入数 + CRM 端数据
    metrics["valid_leads"] = local_valid + crm_res["valid_leads"]
    metrics["potential_leads"] = local_potential + crm_res["potential_leads"]
    metrics["production_value"] = crm_res["production_value"]
    metrics["receive_value"] = crm_res["receive_value"]
    
    return {
        "group_name": group_name,
        "user_ids": user_ids,
        "user_names": user_names,
        "members_list": [(m.id, m.name, m.crm_user_id) for m in members],
        "metrics": metrics,
        "crm_details_text": crm_res["crm_details_text"]
    }


@router.get("/weekly/group-report", response_model=GroupWeeklyReportResponse, summary="读取团队已存整体周报")
async def get_group_weekly_report(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """读取指定周开始日期下，该战队或三级巴已保存的团队周报快照"""
    stmt = select(GroupWeeklyReport).where(
        GroupWeeklyReport.start_date == start_date
    )
    if team_id:
        stmt = stmt.where(GroupWeeklyReport.team_id == team_id)
    else:
        stmt = stmt.where(GroupWeeklyReport.team_id == None)
        
    if third_class_bar and third_class_bar != "all":
        stmt = stmt.where(GroupWeeklyReport.third_class_bar == third_class_bar)
    else:
        stmt = stmt.where(GroupWeeklyReport.third_class_bar == None)
        
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    
    if not report:
        raise HTTPException(status_code=404, detail="未找到该周保存的团队周报")
        
    return report


@router.post("/weekly/save-group-report", response_model=GroupWeeklyReportResponse, summary="保存或覆盖更新团队整体周报")
async def save_group_weekly_report(
    report_in: GroupWeeklyReportSave,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """保存或覆盖更新团队整体周报记录及指标快照"""
    from app.models.user import UserRole
    allowed_roles = [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value, UserRole.TEAM_LEADER.value, "digital_specialist"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足，您无权保存团队整体周报")

    stmt = select(GroupWeeklyReport).where(
        GroupWeeklyReport.start_date == report_in.start_date
    )
    if report_in.team_id:
        stmt = stmt.where(GroupWeeklyReport.team_id == report_in.team_id)
    else:
        stmt = stmt.where(GroupWeeklyReport.team_id == None)
        
    if report_in.third_class_bar and report_in.third_class_bar != "all":
        stmt = stmt.where(GroupWeeklyReport.third_class_bar == report_in.third_class_bar)
    else:
        stmt = stmt.where(GroupWeeklyReport.third_class_bar == None)
        
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    
    if report:
        # 更新现有周报
        report.content = report_in.content
        report.marketing_signed = report_in.marketing_signed
        report.delivery_signed = report_in.delivery_signed
        report.win_bids = report_in.win_bids
        report.happiness_count = report_in.happiness_count
        report.triangle_count = report_in.triangle_count
        report.valid_leads = report_in.valid_leads
        report.potential_leads = report_in.potential_leads
        report.production_value = report_in.production_value
        report.receive_value = report_in.receive_value
        report.updated_at = datetime.now(timezone.utc)
    else:
        # 创建新周报
        report = GroupWeeklyReport(
            team_id=report_in.team_id,
            third_class_bar=report_in.third_class_bar if report_in.third_class_bar != "all" else None,
            start_date=report_in.start_date,
            end_date=report_in.end_date,
            content=report_in.content,
            marketing_signed=report_in.marketing_signed,
            delivery_signed=report_in.delivery_signed,
            win_bids=report_in.win_bids,
            happiness_count=report_in.happiness_count,
            triangle_count=report_in.triangle_count,
            valid_leads=report_in.valid_leads,
            potential_leads=report_in.potential_leads,
            production_value=report_in.production_value,
            receive_value=report_in.receive_value,
            created_by=current_user.id
        )
        db.add(report)
        
    await db.commit()
    await db.refresh(report)
    
    return report


# 记录正在生成的团队周报状态
# key 格式: "team_id:third_class_bar:start_date"
# value 格式: {"status": "running" | "success" | "failed", "error": str | None, "updated_at": datetime}
weekly_report_tasks = {}


async def async_generate_group_report_task(
    start_date: date,
    team_id: int | None,
    third_class_bar: str | None,
    user_id: int,
    task_key: str
):
    """后台异步执行 AI 整体周报生成并自动存盘数据库的任务"""
    from app.database import AsyncSessionLocal
    from datetime import datetime, timezone, timedelta
    from app.models.report import WeeklyReport, GroupWeeklyReport, DailyReport, ReportDetail, ReportStatus
    from app.models.user import User
    from sqlalchemy import select
    import logging
    import re

    logger = logging.getLogger("battle100")
    logger.info(f"开始后台 AI 团队周报生成任务: {task_key}")

    async with AsyncSessionLocal() as db:
        try:
            # 1. 自动提取指标与上下文
            metrics_result = await get_group_weekly_metrics(db, start_date, team_id, third_class_bar)
            
            user_ids = metrics_result["user_ids"]
            if not user_ids:
                raise Exception("该战队或三级巴下暂无激活成员，无法生成周报")
                
            # 2. 收集全员周报与未交日报明细
            weekly_stmt = select(WeeklyReport, User.name).join(User, WeeklyReport.user_id == User.id).where(
                WeeklyReport.user_id.in_(user_ids),
                WeeklyReport.start_date == start_date,
                WeeklyReport.status == "submitted"
            )
            weekly_res = await db.execute(weekly_stmt)
            weekly_rows = weekly_res.all()
            
            submitted_user_ids = set()
            weekly_contents = []
            
            for wr, u_name in weekly_rows:
                submitted_user_ids.add(wr.user_id)
                w_text = (
                    f"--- 成员: {u_name} 的个人周复盘 ---\n"
                    f"【项目交付实际完成】: {wr.delivery_actual or '无'}\n"
                    f"【销售实际完成】: {wr.sales_actual or '无'}\n"
                    f"【本周项目亮点】: {wr.delivery_highlights or '无'}\n"
                    f"【本周销售亮点】: {wr.sales_highlights or '无'}\n"
                    f"【本周项目难点】: {wr.delivery_blockers or '无'}\n"
                    f"【本周销售难点】: {wr.sales_blockers or '无'}\n"
                    f"【下周交付计划】: {wr.next_delivery_plan or '无'}\n"
                    f"【下周销售计划】: {wr.next_sales_plan or '无'}\n"
                )
                weekly_contents.append(w_text)
                
            # 找出未交周报的成员名单
            unsubmitted_members = []
            unsubmitted_user_ids = []
            for uid, uname, _ in metrics_result["members_list"]:
                if uid not in submitted_user_ids:
                    unsubmitted_members.append(uname)
                    unsubmitted_user_ids.append(uid)
                    
            # 对于未提交周报的成员，提取他们本周全量的日报日志明细
            daily_contents = []
            if unsubmitted_user_ids:
                # 查询当周已审核日报及其明细
                daily_stmt = select(DailyReport).where(
                    DailyReport.user_id.in_(unsubmitted_user_ids),
                    DailyReport.report_date >= start_date,
                    DailyReport.report_date <= (start_date + timedelta(days=6)),
                    DailyReport.status == ReportStatus.REVIEWED
                ).order_by(DailyReport.report_date.asc())
                daily_res = await db.execute(daily_stmt)
                daily_reports = daily_res.scalars().all()
                
                user_daily_map = {}
                for dr in daily_reports:
                    if dr.user_id not in user_daily_map:
                        user_daily_map[dr.user_id] = []
                    
                    # 读取该日报的明细
                    detail_stmt = select(ReportDetail).where(ReportDetail.report_id == dr.id)
                    detail_res = await db.execute(detail_stmt)
                    details = detail_res.scalars().all()
                    
                    detail_texts = []
                    for dt in details:
                        detail_texts.append(f"  - [{dt.detail_type}] {dt.project_name or dt.customer_name or ''}: {dt.description or ''}")
                        
                    dr_text = (
                        f"【{dr.report_date.strftime('%m-%d')} 日报】:\n"
                        f"  工作总结: {dr.work_summary or '未填'}\n"
                        f"  核心动作明细:\n" + ("\n".join(detail_texts) if detail_texts else "  无动作明细")
                    )
                    user_daily_map[dr.user_id].append(dr_text)
                    
                for uid, uname, _ in metrics_result["members_list"]:
                    if uid in unsubmitted_user_ids:
                        d_reports_list = user_daily_map.get(uid, [])
                        d_text = (
                            f"--- 成员: {uname} (本周未提交个人周报，以下为其本周全量日报流水) ---\n"
                            + ("\n".join(d_reports_list) if d_reports_list else "本周无已审核的日报动作") + "\n"
                        )
                        daily_contents.append(d_text)
                        
            # 3. 构造给 AI 汇总的数据文本
            team_name_title = metrics_result["group_name"]
            metrics = metrics_result["metrics"]
            
            metrics_summary_text = (
                f"【本周 {team_name_title} 核心汇总业绩数据看板】:\n"
                f"- 营销新签合同额: {metrics['marketing_signed']:.2f} 万元\n"
                f"- 交付新签合同额: {metrics['delivery_signed']:.2f} 万元\n"
                f"- 中标项目个数: {metrics['win_bids']} 个\n"
                f"- 幸福动作个数: {metrics['happiness_count']} 次\n"
                f"- 铁三角联动次数: {metrics['triangle_count']} 次\n"
                f"- 有效商机线索量: {metrics['valid_leads']} 个\n"
                f"- 潜力商机线索量: {metrics['potential_leads']} 个\n"
                f"- CRM 累计产值: {metrics['production_value']:.2f} 万元\n"
                f"- CRM 到账回款额: {metrics['receive_value']:.2f} 万元\n"
            )
            
            crm_details = metrics_result.get("crm_details_text", "暂无 CRM 沟通或项目跟进明细")
            
            report_data_context = (
                f"目前正在为团队【{team_name_title}】生成本周的整体复盘周报。\n"
                f"时间跨度为: {start_date} ~ {start_date + timedelta(days=6)}\n\n"
                f"{metrics_summary_text}\n"
                f"【团队 CRM 业务明细跟进与沟通原文】:\n{crm_details}\n\n"
                f"【已提交周报的成员复盘详情】:\n" + ("\n".join(weekly_contents) if weekly_contents else "无已提交的个人周报。\n") + "\n"
                f"【未提交周报的成员全量已审核日报描述】:\n" + ("\n".join(daily_contents) if daily_contents else "无未提交周报的成员日报数据。\n") + "\n"
            )
            
            unsubmitted_str = "、".join(unsubmitted_members) if unsubmitted_members else "无"
            
            # 4. 调用大模型路由进行周报编写
            from app.llm.provider import get_provider_for_agent
            from app.models.llm_config import AgentRoute
            
            provider, model_id = await get_provider_for_agent("reports")
            
            route_stmt = select(AgentRoute).where(AgentRoute.agent_role == "reports")
            route_res = await db.execute(route_stmt)
            route_obj = route_res.scalar_one_or_none()
            
            default_sys_prompt = (
                "你是“百日奋战”战队及三级巴整体复盘的【战报及铁三角联动分析师】。\n"
                "你的任务是根据提供的团队多维汇总数据（包括营销新签、交付新签、中标数、幸福动作、铁三角联动、CRM 产值、CRM 到账回款）以及每个成员的个人周报/日报工作流水，撰写并整理出一份精美、专业、条理清晰的【团队整体周报】。\n"
                "要求：\n"
                "1. 整体周报采用 Markdown 格式输出。内容必须切合数据，严禁虚构、夸大事实，突出团队的签约、回款以及里程碑交付成果。\n"
                "2. 语言风格要充满战斗激情、逻辑严密，并提出下阶段重点攻坚建议。\n"
                "3. 报告必须包含以下板块：\n"
                "   - **一、团队本周核心业绩看板**（用 Markdown 表格展示提供给你的所有汇总指标快照，包括产值、回款、签约等）；\n"
                "   - **二、本周工作主要战果与亮点**（结合个人周报/日报及 CRM 沟通纪要，细致描述具体项目的推进、突破、大额签约或回款情况，提及具体负责人）；\n"
                "   - **三、交付卡点与重大业务预警**：\n"
                "     1. 深度分析团队面临的暂停/延期项目，特别是‘已到交付节点未开票’和‘已开票未回款’的项目；\n"
                "     2. **必须依据【正在实施项目最新进度清单】与【项目进度异动记录】对所有团队成员的工作饱和度及项目健康度进行诊断，并以专门的子标题加粗高亮输出以下警报与提示：**\n"
                "        - 🚨 **红色警报（无项目人员）**：若有成员名下没有任何当前活跃正在实施的项目（在清单中未出现），必须高亮列出，形式如：`🚨 红色警报：[姓名] 名下无任何正在实施项目，需立即核实并安排任务！`；\n"
                "        - ⚠️ **黄色预警（项目停滞人员）**：若有成员名下有项目，但该项目本周没有任何进度异动推进（即进度变化为0%且无任何日报推进说明），必须高亮列出，形式如：`⚠️ 黄色预警：[姓名] 负责的项目本周进度停滞，无任何推进，需重点关注！`；\n"
                "        - 💡 **风险提示（项目空仓人员）**：若有成员名下负责的项目极少（≤ 2 个）且所有正在实施项目的当前最新进度均已接近完成（最新进度均已 ≥ 90%，包含 100% 已完工的项目），说明其即将面临‘断粮’空仓风险，必须高亮列出，形式如：`💡 风险提示：[姓名] 名下仅有 [X] 个正在实施项目且均已接近完成（当前进度≥90%），面临项目断档风险，需尽快规划新项目接入。`；\n"
                "        - **注意**：以上警报及提示项下，【仅列出符合该预警条件的成员】。正常推进、饱和度良好的成员【绝对不要】列入上述任何异常警报列表中。如果没有任何人符合某警报条件，则直接在该警报板块下输出肯定结论（如‘✅ 经核查，全员本周均有正在实施的活跃项目，无无项目红色警报。’）。\n"
                "   - **四、下周重点攻坚方向与计划**（结合个人下周目标给出团队攻坚计划）；\n"
                "   - **五、需要协调与支持的事项**（明确指出当前阻碍交付或签约、回款的卡点，并点明需要上级或跨部门支持的具体事项）。\n"
                f"4. 你必须在 Markdown 文本的最后，用专门的一行渲染 `【本周未填报人员】：{unsubmitted_str}`。"
            )
            
            db_system_prompt = route_obj.system_prompt if route_obj else None
            db_user_prompt = route_obj.user_prompt if route_obj else None
            
            if route_obj and route_obj.system_prompt and ("仅列出符合该预警条件的成员" not in route_obj.system_prompt or "在研项目" in route_obj.system_prompt):
                route_obj.system_prompt = default_sys_prompt
                db.add(route_obj)
                await db.commit()
                await db.refresh(route_obj)
                db_system_prompt = route_obj.system_prompt
                
            system_prompt = db_system_prompt if db_system_prompt else default_sys_prompt
            if db_user_prompt:
                try:
                    user_prompt = db_user_prompt.format(report_data=report_data_context)
                except Exception:
                    user_prompt = db_user_prompt + f"\n\n{report_data_context}"
            else:
                user_prompt = f"请为我们战队/三级巴撰写本周的整体分析周报，以下是团队的各项业绩指标 and 全员本周记录详情：\n\n{report_data_context}"
                
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ]
            
            # 增加最大输出限制到 16384 个 Token，防止超长团队周报被中途截断，所有注释必须使用中文
            ai_content = await provider.chat(messages, max_tokens=16384)
            ai_content = re.sub(r"<think>.*?</think>", "", ai_content, flags=re.DOTALL).strip()

            # 5. 自动保存到数据库
            stmt = select(GroupWeeklyReport).where(
                GroupWeeklyReport.start_date == start_date
            )
            if team_id:
                stmt = stmt.where(GroupWeeklyReport.team_id == team_id)
            else:
                stmt = stmt.where(GroupWeeklyReport.team_id == None)
                
            if third_class_bar and third_class_bar != "all":
                stmt = stmt.where(GroupWeeklyReport.third_class_bar == third_class_bar)
            else:
                stmt = stmt.where(GroupWeeklyReport.third_class_bar == None)
                
            res = await db.execute(stmt)
            report = res.scalar_one_or_none()
            
            if report:
                report.content = ai_content
                report.marketing_signed = metrics.get('marketing_signed', 0.0)
                report.delivery_signed = metrics.get('delivery_signed', 0.0)
                report.win_bids = metrics.get('win_bids', 0)
                report.happiness_count = metrics.get('happiness_count', 0)
                report.triangle_count = metrics.get('triangle_count', 0)
                report.valid_leads = metrics.get('valid_leads', 0)
                report.potential_leads = metrics.get('potential_leads', 0)
                report.production_value = metrics.get('production_value', 0.0)
                report.receive_value = metrics.get('receive_value', 0.0)
                report.updated_at = datetime.now(timezone.utc)
            else:
                report = GroupWeeklyReport(
                    team_id=team_id,
                    third_class_bar=third_class_bar if third_class_bar != "all" else None,
                    start_date=start_date,
                    end_date=start_date + timedelta(days=6),
                    content=ai_content,
                    marketing_signed=metrics.get('marketing_signed', 0.0),
                    delivery_signed=metrics.get('delivery_signed', 0.0),
                    win_bids=metrics.get('win_bids', 0),
                    happiness_count=metrics.get('happiness_count', 0),
                    triangle_count=metrics.get('triangle_count', 0),
                    valid_leads=metrics.get('valid_leads', 0),
                    potential_leads=metrics.get('potential_leads', 0),
                    production_value=metrics.get('production_value', 0.0),
                    receive_value=metrics.get('receive_value', 0.0),
                    created_by=user_id
                )
                db.add(report)
                
            await db.commit()
            
            # 更新状态为 success
            weekly_report_tasks[task_key] = {
                "status": "success",
                "error": None,
                "updated_at": datetime.now(timezone.utc)
            }
            logger.info(f"后台 AI 团队周报生成任务成功: {task_key}")
            
        except Exception as e:
            weekly_report_tasks[task_key] = {
                "status": "failed",
                "error": str(e),
                "updated_at": datetime.now(timezone.utc)
            }
            logger.error(f"后台 AI 团队周报生成任务失败: {task_key}, 错误: {e}")


@router.get("/weekly/generate-status", summary="查询 AI 智能生成状态")
async def get_generate_status(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    current_user: User = Depends(get_current_user),
):
    """查询指定战队或三级巴在当周的 AI 智能生成后台任务状态"""
    from app.models.user import UserRole
    allowed_roles = [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value, UserRole.TEAM_LEADER.value, "digital_specialist"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足")
        
    t_str = str(team_id) if team_id else "None"
    b_str = third_class_bar if third_class_bar and third_class_bar != "all" else "None"
    task_key = f"{t_str}:{b_str}:{start_date.strftime('%Y-%m-%d')}"
    
    task = weekly_report_tasks.get(task_key)
    if not task:
        return {"status": "idle", "error": None}
        
    # 如果任务完成时间超过 10 分钟，清理掉（返回 idle 并且从字典删除），避免残留
    if task["status"] in ["success", "failed"]:
        task_time = task["updated_at"]
        if (datetime.now(timezone.utc) - task_time).total_seconds() > 600:
            weekly_report_tasks.pop(task_key, None)
            return {"status": "idle", "error": None}
            
    return {
        "status": task["status"],
        "error": task["error"]
    }


@router.post("/weekly/generate-group-report", summary="AI 智能生成团队整体周报")
async def generate_group_weekly_report(
    background_tasks: BackgroundTasks,
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """在后台异步启动 AI 生成战队/三级巴整体周报并自动存盘"""
    from app.models.user import UserRole
    
    allowed_roles = [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value, UserRole.TEAM_LEADER.value, "digital_specialist"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足，您无权生成团队整体周报")
        
    t_str = str(team_id) if team_id else "None"
    b_str = third_class_bar if third_class_bar and third_class_bar != "all" else "None"
    task_key = f"{t_str}:{b_str}:{start_date.strftime('%Y-%m-%d')}"
    
    # 检查当前是否已经在生成中
    if task_key in weekly_report_tasks and weekly_report_tasks[task_key]["status"] == "running":
        task_time = weekly_report_tasks[task_key]["updated_at"]
        # 超过 5 分钟的超时防呆设计，允许重新生成
        if (datetime.now(timezone.utc) - task_time).total_seconds() < 300:
            raise HTTPException(status_code=400, detail="该团队当周的 AI 周报正由后台整理生成中，请勿重复触发！")
            
    # 初始化状态为 running
    weekly_report_tasks[task_key] = {
        "status": "running",
        "error": None,
        "updated_at": datetime.now(timezone.utc)
    }
    
    # 放入后台执行
    background_tasks.add_task(
        async_generate_group_report_task,
        start_date,
        team_id,
        third_class_bar,
        current_user.id,
        task_key
    )
    
    return {
        "status": "processing",
        "task_key": task_key,
        "message": "AI 团队整体周报正在后台汇总整理中，预计需要 1-2 分钟，您可以继续处理其他工作。生成完毕后数据库将自动存盘。"
    }


# 下面这个是已经废弃的同步生成逻辑，重命名防止路由冲突与编译错误
async def old_generate_group_weekly_report(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """根据本组数据与 CRM 产值、周报/日报明细，自动驱动 AI 生成战队/三级巴整体周报 (放宽至 20K-30K 上下文)"""
    from app.models.user import UserRole
    from app.models.report import WeeklyReport
    from datetime import timedelta
    
    allowed_roles = [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value, UserRole.TEAM_LEADER.value, "digital_specialist"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足，您无权生成团队整体周报")
        
    # 1. 自动提取指标与上下文
    metrics_result = await get_group_weekly_metrics(db, start_date, team_id, third_class_bar)
    
    user_ids = metrics_result["user_ids"]
    if not user_ids:
        raise HTTPException(status_code=400, detail="该战队或三级巴下暂无激活成员，无法生成周报")
        
    # 2. 收集全员周报与未交日报明细
    weekly_stmt = select(WeeklyReport, User.name).join(User, WeeklyReport.user_id == User.id).where(
        WeeklyReport.user_id.in_(user_ids),
        WeeklyReport.start_date == start_date,
        WeeklyReport.status == "submitted"
    )
    weekly_res = await db.execute(weekly_stmt)
    weekly_rows = weekly_res.all()
    
    submitted_user_ids = set()
    weekly_contents = []
    
    for wr, u_name in weekly_rows:
        submitted_user_ids.add(wr.user_id)
        w_text = (
            f"--- 成员: {u_name} 的个人周复盘 ---\n"
            f"【项目交付实际完成】: {wr.delivery_actual or '无'}\n"
            f"【销售实际完成】: {wr.sales_actual or '无'}\n"
            f"【本周项目亮点】: {wr.delivery_highlights or '无'}\n"
            f"【本周销售亮点】: {wr.sales_highlights or '无'}\n"
            f"【本周项目难点】: {wr.delivery_blockers or '无'}\n"
            f"【本周销售难点】: {wr.sales_blockers or '无'}\n"
            f"【下周交付计划】: {wr.next_delivery_plan or '无'}\n"
            f"【下周销售计划】: {wr.next_sales_plan or '无'}\n"
        )
        weekly_contents.append(w_text)
        
    # 找出未交周报的成员名单
    unsubmitted_members = []
    unsubmitted_user_ids = []
    for uid, uname, _ in metrics_result["members_list"]:
        if uid not in submitted_user_ids:
            unsubmitted_members.append(uname)
            unsubmitted_user_ids.append(uid)
            
    # 对于未提交周报的成员，提取他们本周全量的日报日志明细，放宽上下文到 20K-30K 传全量日报描述
    daily_contents = []
    if unsubmitted_user_ids:
        # 查询当周已审核日报及其明细
        daily_stmt = select(DailyReport).where(
            DailyReport.user_id.in_(unsubmitted_user_ids),
            DailyReport.report_date >= start_date,
            DailyReport.report_date <= (start_date + timedelta(days=6)),
            DailyReport.status == ReportStatus.REVIEWED
        ).order_by(DailyReport.report_date.asc())
        daily_res = await db.execute(daily_stmt)
        daily_reports = daily_res.scalars().all()
        
        user_daily_map = {}
        for dr in daily_reports:
            if dr.user_id not in user_daily_map:
                user_daily_map[dr.user_id] = []
            
            # 读取该日报的明细
            detail_stmt = select(ReportDetail).where(ReportDetail.report_id == dr.id)
            detail_res = await db.execute(detail_stmt)
            details = detail_res.scalars().all()
            
            detail_texts = []
            for dt in details:
                detail_texts.append(f"  - [{dt.detail_type}] {dt.project_name or dt.customer_name or ''}: {dt.description or ''}")
                
            dr_text = (
                f"【{dr.report_date.strftime('%m-%d')} 日报】:\n"
                f"  工作总结: {dr.work_summary or '未填'}\n"
                f"  核心动作明细:\n" + ("\n".join(detail_texts) if detail_texts else "  无动作明细")
            )
            user_daily_map[dr.user_id].append(dr_text)
            
        for uid, uname, _ in metrics_result["members_list"]:
            if uid in unsubmitted_user_ids:
                d_reports_list = user_daily_map.get(uid, [])
                d_text = (
                    f"--- 成员: {uname} (本周未提交个人周报，以下为其本周全量日报流水) ---\n"
                    + ("\n".join(d_reports_list) if d_reports_list else "本周无已审核的日报动作") + "\n"
                )
                daily_contents.append(d_text)
                
    # 3. 构造给 AI 汇总的数据文本
    team_name_title = metrics_result["group_name"]
    metrics = metrics_result["metrics"]
    
    metrics_summary_text = (
        f"【本周 {team_name_title} 核心汇总业绩数据看板】:\n"
        f"- 营销新签合同额: {metrics['marketing_signed']:.2f} 万元\n"
        f"- 交付新签合同额: {metrics['delivery_signed']:.2f} 万元\n"
        f"- 中标项目个数: {metrics['win_bids']} 个\n"
        f"- 幸福动作个数: {metrics['happiness_count']} 次\n"
        f"- 铁三角联动次数: {metrics['triangle_count']} 次\n"
        f"- 有效商机线索量: {metrics['valid_leads']} 个\n"
        f"- 潜力商机线索量: {metrics['potential_leads']} 个\n"
        f"- CRM 累计产值: {metrics['production_value']:.2f} 万元\n"
        f"- CRM 到账回款额: {metrics['receive_value']:.2f} 万元\n"
    )
    
    crm_details = metrics_result.get("crm_details_text", "暂无 CRM 沟通或项目跟进明细")
    
    report_data_context = (
        f"目前正在为团队【{team_name_title}】生成本周的整体复盘周报。\n"
        f"时间跨度为: {start_date} ~ {start_date + timedelta(days=6)}\n\n"
        f"{metrics_summary_text}\n"
        f"【团队 CRM 业务明细跟进与沟通原文】:\n{crm_details}\n\n"
        f"【已提交周报的成员复盘详情】:\n" + ("\n".join(weekly_contents) if weekly_contents else "无已提交的个人周报。\n") + "\n"
        f"【未提交周报的成员全量已审核日报描述】:\n" + ("\n".join(daily_contents) if daily_contents else "无未提交周报的成员日报数据。\n") + "\n"
    )
    
    unsubmitted_str = "、".join(unsubmitted_members) if unsubmitted_members else "无"
    
    # 4. 调用大模型路由进行周报编写
    try:
        from app.llm.provider import get_provider_for_agent
        from app.models.llm_config import AgentRoute
        
        provider, model_id = await get_provider_for_agent("reports")
        
        # 优先读取数据库配置的自定义 system_prompt 与 user_prompt 模板
        route_stmt = select(AgentRoute).where(AgentRoute.agent_role == "reports")
        route_res = await db.execute(route_stmt)
        route_obj = route_res.scalar_one_or_none()
        
        # 默认强力 System Prompt 作为兜底
        default_sys_prompt = (
            "你是“百日奋战”战队及三级巴整体复盘的【战报及铁三角联动分析师】。\n"
            "你的任务是根据提供的团队多维汇总数据（包括营销新签、交付新签、中标数、幸福动作、铁三角联动、CRM 产值、CRM 到账回款）以及每个成员的个人周报/日报工作流水，撰写并整理出一份精美、专业、条理清晰的【团队整体周报】。\n"
            "要求：\n"
            "1. 整体周报采用 Markdown 格式输出。内容必须切合数据，严禁虚构、夸大事实，突出团队的签约、回款以及里程碑交付成果。\n"
            "2. 语言风格要充满战斗激情、逻辑严密，并提出下阶段重点攻坚建议。\n"
            "3. 报告必须包含以下板块：\n"
            "   - **一、团队本周核心业绩看板**（用 Markdown 表格展示提供给你的所有汇总指标快照，包括产值、回款、签约等）；\n"
            "   - **二、本周工作主要战果与亮点**（结合个人周报/日报及 CRM 沟通纪要，细致描述具体项目的推进、突破、大额签约或回款情况，提及具体负责人）；\n"
            "   - **三、交付卡点与重大业务预警**：\n"
            "     1. 深度分析团队面临的暂停/延期项目，特别是‘已到交付节点未开票’和‘已开票未回款’的项目；\n"
            "     2. **必须依据【正在实施项目最新进度清单】与【项目进度异动记录】对所有团队成员的工作饱和度及项目健康度进行诊断，并以专门的子标题加粗高亮输出以下警报与提示：**\n"
            "        - 🚨 **红色警报（无项目人员）**：若有成员名下没有任何当前活跃正在实施的项目（在清单中未出现），必须高亮列出，形式如：`🚨 红色警报：[姓名] 名下无任何正在实施项目，需立即核实并安排任务！`；\n"
            "        - ⚠️ **黄色预警（项目停滞人员）**：若有成员名下有项目，但该项目本周没有任何进度异动推进（即进度变化为0%且无任何日报推进说明），必须高亮列出，形式如：`⚠️ 黄色预警：[姓名] 负责的项目本周进度停滞，无任何推进，需重点关注！`；\n"
            "        - 💡 **风险提示（项目空仓人员）**：若有成员名下负责的项目极少（≤ 2 个）且所有正在实施项目的当前最新进度均已接近完成（最新进度均已 ≥ 90%，包含 100% 已完工的项目），说明其即将面临‘断粮’空仓风险，必须高亮列出，形式如：`💡 风险提示：[姓名] 名下仅有 [X] 个正在实施项目且均已接近完成（当前进度≥90%），面临项目断档风险，需尽快规划新项目接入。`；\n"
            "        - **注意**：以上警报及提示项下，【仅列出符合该预警条件的成员】。正常推进、饱和度良好的成员【绝对不要】列入上述任何异常警报列表中。如果没有任何人符合某警报条件，则直接在该警报板块下输出肯定结论（如‘✅ 经核查，全员本周均有正在实施的活跃项目，无无项目红色警报。’）。\n"
            "   - **四、下周重点攻坚方向与计划**（结合个人下周目标给出团队攻坚计划）；\n"
            "   - **五、需要协调与支持的事项**（明确指出当前阻碍交付或签约、回款的卡点，并点明需要上级或跨部门支持的具体事项）。\n"
            f"4. 你必须在 Markdown 文本的最后，用专门的一行渲染 `【本周未填报人员】：{unsubmitted_str}`。"
        )

        db_system_prompt = route_obj.system_prompt if route_obj else None
        db_user_prompt = route_obj.user_prompt if route_obj else None
        
        # 若数据库中已有配置，但尚未更新成员工作饱和度诊断逻辑，或提示词中仍包含旧的“在研项目”字样，自动重设更新数据库中的 system_prompt 确保百分之百生效
        if route_obj and route_obj.system_prompt and ("仅列出符合该预警条件的成员" not in route_obj.system_prompt or "在研项目" in route_obj.system_prompt):
            route_obj.system_prompt = default_sys_prompt
            db.add(route_obj)
            await db.commit()
            await db.refresh(route_obj)
            db_system_prompt = route_obj.system_prompt

        # 确定最终 prompt
        system_prompt = db_system_prompt if db_system_prompt else default_sys_prompt
        # 对用户提示词，如果系统配了，则以系统配的模板把 {report_data} 替换掉
        if db_user_prompt:
            try:
                user_prompt = db_user_prompt.format(report_data=report_data_context)
            except Exception:
                user_prompt = db_user_prompt + f"\n\n{report_data_context}"
        else:
            user_prompt = f"请为我们战队/三级巴撰写本周的整体分析周报，以下是团队的各项业绩指标 and 全员本周记录详情：\n\n{report_data_context}"
            
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
        # 增加最大输出限制到 16384 个 Token，防止超长团队周报（如多项目大表格及思考内容）被中途截断，所有注释必须使用中文
        ai_content = await provider.chat(messages, max_tokens=16384)
        
        # 自动清洗掉思考过程（针对思考类 LLM 模型输出的 <think>...</think>）
        import re
        ai_content = re.sub(r"<think>.*?</think>", "", ai_content, flags=re.DOTALL).strip()
        
    except Exception as e:
        import logging
        logging.getLogger("battle100").error(f"AI 生成团队周报出错: {e}")
        ai_content = f"### AI 生成团队周报失败\n\n错误信息：{str(e)}\n\n请尝试点击“重新由 AI 智能生成”。"
        
    return {
        "metrics": metrics,
        "content": ai_content,
        "unsubmitted_members": unsubmitted_members
    }


class SendGroupReportToDingtalkRequest(BaseModel):
    group_name: str
    start_date: date
    metrics: dict
    content: str
    redirect_url: str | None = None


@router.post("/weekly/send-group-report-to-dingtalk", summary="推送整体周报到钉钉群")
async def send_group_report_to_dingtalk_api(
    req: SendGroupReportToDingtalkRequest,
    current_user: User = Depends(get_current_user),
):
    """人工点击一键推送整体周报到钉钉机器人"""
    from app.models.user import UserRole
    allowed_roles = [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value, UserRole.TEAM_LEADER.value, "digital_specialist"]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足，您无权向钉钉推送团队整体周报")
        
    from app.services.dingtalk import send_group_weekly_report_to_dingtalk
    success = await send_group_weekly_report_to_dingtalk(
        group_name=req.group_name,
        start_date_val=req.start_date,
        metrics=req.metrics,
        content=req.content,
        redirect_url=req.redirect_url
    )
    
    if not success:
        raise HTTPException(status_code=500, detail="推送至钉钉机器人失败，请检查配置或稍后重试")
        
    return {"message": "已成功推送至钉钉群机器人"}


class ExportDocxRequest(BaseModel):
    title: str
    metrics: dict | None = None
    content: str


@router.post("/weekly/export-docx", summary="导出周报为 Word (.docx) 文件")
async def export_weekly_report_to_docx_api(
    req: ExportDocxRequest,
    current_user: User = Depends(get_current_user)
):
    """
    接收 Markdown 周报正文及指标看板数据，在后端动态生成排版精美的 Word 二进制文档并输出下载流
    """
    from app.services.docx_exporter import export_markdown_to_docx
    from fastapi.responses import StreamingResponse
    import urllib.parse
    
    try:
        # 1. 转换并生成 Word 字节流
        file_stream = export_markdown_to_docx(
            title=req.title,
            metrics=req.metrics,
            content=req.content
        )
        
        # 2. 对文件名进行 URL 编码以支持中文字符集
        safe_filename = urllib.parse.quote(f"{req.title}.docx")
        
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers=headers
        )
    except Exception as e:
        import logging
        logging.getLogger("battle100").error(f"导出 Word 文档出错: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Word 导出失败: {str(e)}")


@router.get("/{report_id}", response_model=DailyReportResponse, summary="获取填报详情")
async def get_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取指定填报的详细信息"""
    result = await db.execute(
        select(DailyReport)
        .options(selectinload(DailyReport.details))
        .where(DailyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=404, detail="填报记录不存在")

    # 权限检查
    if current_user.role == UserRole.STAFF.value and report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权查看他人的填报记录")

    return report


class SyncToDingTalkRequest(BaseModel):
    template_id: Optional[str] = None


@router.post("/weekly/{report_id}/sync-to-dingtalk", summary="同步周报到钉钉工作日志")
async def sync_weekly_report_to_dingtalk_api(
    report_id: int,
    req: SyncToDingTalkRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    一键同步本系统的周报到钉钉工作日志，所有注释必须使用中文
    """
    from app.models.report import WeeklyReport
    from app.models.user import User as DbUser
    from app.integrations.dingtalk import dingtalk_client
    from app.config import settings

    # 1. 查找并校验周报记录
    stmt = select(WeeklyReport).where(WeeklyReport.id == report_id)
    res = await db.execute(stmt)
    report = res.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="未找到对应的周报记录")

    # 2. 权限校验，只能同步自己的周报
    if report.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="您无权同步他人的周报记录")

    # 3. 获取当前用户的详细记录，提前查询以避免 NameError，所有注释必须使用中文
    db_user_stmt = select(DbUser).where(DbUser.id == current_user.id)
    db_user_res = await db.execute(db_user_stmt)
    db_user = db_user_res.scalar_one_or_none()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 4. 模板 ID 获取与专属模板强制重定向逻辑，所有注释必须使用中文
    custom_template_id = None
    team_name = ""
    if db_user.team_id:
        from app.models.organization import Team
        team_obj = await db.get(Team, db_user.team_id)
        if team_obj:
            team_name = team_obj.name
            
    # 尝试从环境变量读取战队专属的 Webhook JSON 配置，识别战队专属模板
    import json
    import os
    team_webhooks = {}
    env_config = os.getenv("TEAM_WEBHOOKS_JSON")
    if not env_config:
        # 防御性读取：若系统未能成功将多行 JSON 加载至环境变量，主动从当前目录下的 .env 文件再次加载，所有注释必须使用中文
        try:
            from dotenv import load_dotenv
            load_dotenv(r"c:\APP\AI100\battle100\backend\.env")
            env_config = os.getenv("TEAM_WEBHOOKS_JSON")
        except Exception:
            pass

    if env_config:
        try:
            team_webhooks = json.loads(env_config)
        except Exception as e_cfg:
            pass
            
    if team_name and team_name in team_webhooks:
        custom_cfg = team_webhooks[team_name]
        if custom_cfg.get("template_id"):
            custom_template_id = custom_cfg["template_id"]

    # 确定最终使用的模板 ID：如果配置了战队专属模板，则强制重定向使用，确保流向正确
    if custom_template_id:
        template_id = custom_template_id
    else:
        # 如果未配置战队专属模板，则使用前端传过来的模板 ID；若前端也未传，使用默认大盘模板 ID
        template_id = req.template_id
        if not template_id:
            template_id = settings.DINGTALK_WEEKLY_REPORT_TEMPLATE_ID

    if not template_id:
        raise HTTPException(status_code=400, detail="未指定钉钉模板ID，且系统未配置默认模板ID")
        
    # 自动修正拼写错误的模板ID，解决员工复制或手工填写错误问题，所有注释必须使用中文
    if template_id == "19cab0d8aa4c349cb1df85146edac9cf":
        template_id = "19eab0d8aa4e349cb1df85146edac9cf"

    # 在控制台打印大段显眼的调试日志，方便直接排查和核对运行中的实例，所有注释必须使用中文
    print("\n" + "="*60)
    print(f"【调试周报同步】收到同步请求，报告ID: {report_id}")
    print(f"  员工姓名: {db_user.name}")
    print(f"  所属战队: {team_name} (team_id={db_user.team_id})")
    print(f"  加载的 TEAM_WEBHOOKS_JSON 长度: {len(env_config) if env_config else 0} 字符")
    print(f"  解析得到的配置战队数: {len(team_webhooks) if team_webhooks else 0}")
    print(f"  匹配到的专属模板 ID: '{custom_template_id}'")
    print(f"  前端传入的模板 ID: '{req.template_id}'")
    print(f"  最终决定使用的模板 ID: '{template_id}'")
    print("="*60 + "\n")

    # 5. 获取当前用户的钉钉 userid（工号）
    dingtalk_userid = db_user.dingtalk_id
    if not dingtalk_userid:
        # 如果钉钉userid为空，尝试通过手机号在钉钉匹配获取，实现无感绑定
        dingtalk_userid = await dingtalk_client.get_user_by_mobile(db_user.phone)
        if dingtalk_userid:
            db_user.dingtalk_id = dingtalk_userid
            await db.commit()
            await db.refresh(db_user)
        else:
            raise HTTPException(
                status_code=400, 
                detail="无法匹配您的钉钉身份，请确保系统内注册手机号与企业钉钉所绑定的手机号一致，或请先在个人中心绑定钉钉工号"
            )

    # 5. 按照中台/幸福委周报模板的7个字段中文字样，映射组装表单内容
    is_marketing = db_user.position_type == "marketing"

    plan_val = report.sales_plan if is_marketing else report.delivery_plan
    actual_val = report.sales_actual if is_marketing else report.delivery_actual
    rate_val = report.sales_rate if is_marketing else report.delivery_rate
    highlights_val = report.sales_highlights if is_marketing else report.delivery_highlights
    blockers_val = report.sales_blockers if is_marketing else report.delivery_blockers
    support_val = report.sales_support if is_marketing else report.delivery_support
    next_plan_val = report.next_sales_plan if is_marketing else report.next_delivery_plan

    # 格式化周报日期范围，所有注释必须使用中文
    start_date_str = report.start_date.strftime('%Y-%m-%d') if hasattr(report.start_date, 'strftime') else str(report.start_date)
    end_date_str = report.end_date.strftime('%Y-%m-%d') if hasattr(report.end_date, 'strftime') else str(report.end_date)
    date_range_str = f"{start_date_str}至{end_date_str}"

    contents = [
        {"key": "周报日期", "value": date_range_str, "type": "text"},
        {"key": "本周目标计划", "value": plan_val or "", "type": "text"},
        {"key": "本周实际完成", "value": actual_val or "", "type": "text"},
        {"key": "达成情况", "value": rate_val or "", "type": "text"},
        {"key": "本周亮点", "value": highlights_val or "", "type": "text"},
        {"key": "本周卡点", "value": blockers_val or "", "type": "text"},
        {"key": "是否需要上级支持", "value": support_val or "", "type": "text"},
        {"key": "下周目标", "value": next_plan_val or "", "type": "text"}
    ]

    # 6. 计算日志接收人 (to_userids)，所有注释必须使用中文
    to_userids = []
    
    # 6.1 加入当前操作人（触发同步的人）的钉钉ID
    if current_user.dingtalk_id:
        to_userids.append(current_user.dingtalk_id)
        
    # 6.2 自动获取系统内所有超级管理员的钉钉ID并加入
    stmt_admins = select(DbUser.dingtalk_id).where(
        DbUser.role == "admin",
        DbUser.dingtalk_id.isnot(None),
        DbUser.dingtalk_id != ""
    )
    res_admins = await db.execute(stmt_admins)
    admin_ids = res_admins.scalars().all()
    to_userids.extend(admin_ids)
    
    # 6.3 按照三级巴和战队优先级逻辑，拉取队友的钉钉ID并加入
    # 如果有三级巴设定，则只添加三级巴的队友作为接收人，不再包含战队其他成员。如果没有三级巴，才把整个战队的队友作为接收人。所有注释必须使用中文。
    has_third_bar = db_user.third_class_bar and db_user.third_class_bar.strip() != "" and db_user.third_class_bar.strip().lower() != "all"
    
    if has_third_bar:
        # 6.3.1 若有三级巴，仅将同属该三级巴的队友设为接收人
        stmt_teammates = select(DbUser.dingtalk_id).where(
            DbUser.third_class_bar == db_user.third_class_bar.strip(),
            DbUser.dingtalk_id.isnot(None),
            DbUser.dingtalk_id != ""
        )
        res_teammates = await db.execute(stmt_teammates)
        teammate_ids = res_teammates.scalars().all()
        to_userids.extend(teammate_ids)
    elif db_user.team_id:
        # 6.3.2 若无三级巴但有所属战队，则将该战队内的所有队友（含战队长、专员等）设为接收人
        stmt_teammates = select(DbUser.dingtalk_id).where(
            DbUser.team_id == db_user.team_id,
            DbUser.dingtalk_id.isnot(None),
            DbUser.dingtalk_id != ""
        )
        res_teammates = await db.execute(stmt_teammates)
        teammate_ids = res_teammates.scalars().all()
        to_userids.extend(teammate_ids)
                
    # 6.4 过滤空值，进行去重，且必须排除发件人（创建人）自己，因为钉钉要求接收人列表中不能含有日志创建人
    to_userids = list(set([uid for uid in to_userids if uid and uid != dingtalk_userid]))

    # 7. 调用钉钉工作日志填报 API 填报数据
    success, err_msg = await dingtalk_client.save_report(template_id, dingtalk_userid, contents, to_userids=to_userids)
    if not success:
        raise HTTPException(status_code=500, detail=err_msg)

    return {"message": "已成功同步填报至您的钉钉工作日志"}


async def auto_sync_weekly_report_to_dingtalk_task(report_id: int, user_id: int):
    """
    自动同步周报到钉钉工作日志的后台任务，免去用户手动点击，并自动触发群日志卡片。
    所有注释必须使用中文。
    """
    from app.database import AsyncSessionLocal
    from app.models.report import WeeklyReport
    from app.models.user import User as DbUser
    from app.integrations.dingtalk import dingtalk_client
    from app.config import settings
    import json
    import os
    import logging
    from sqlalchemy import select

    logger = logging.getLogger("battle100.reports.auto_sync")

    async with AsyncSessionLocal() as db:
        try:
            # 1. 查找周报记录
            stmt = select(WeeklyReport).where(WeeklyReport.id == report_id)
            res = await db.execute(stmt)
            report = res.scalar_one_or_none()
            if not report:
                logger.error(f"自动同步周报到钉钉失败: 未找到 ID={report_id} 的周报")
                return

            # 2. 获取用户
            db_user_stmt = select(DbUser).where(DbUser.id == user_id)
            db_user_res = await db.execute(db_user_stmt)
            db_user = db_user_res.scalar_one_or_none()
            if not db_user:
                logger.error(f"自动同步周报到钉钉失败: 未找到 ID={user_id} 的用户")
                return

            # 3. 确定模板 ID
            template_id = None
            team_name = ""
            if db_user.team_id:
                from app.models.organization import Team
                team_obj = await db.get(Team, db_user.team_id)
                if team_obj:
                    team_name = team_obj.name

            team_webhooks = {}
            env_config = os.getenv("TEAM_WEBHOOKS_JSON")
            if not env_config:
                # 防御性读取：若系统未能成功将多行 JSON 加载至环境变量，自动从当前目录下的 .env 文件再次加载，所有注释必须使用中文
                try:
                    from dotenv import load_dotenv
                    load_dotenv(r"c:\APP\AI100\battle100\backend\.env")
                    env_config = os.getenv("TEAM_WEBHOOKS_JSON")
                except Exception:
                    pass

            if env_config:
                try:
                    team_webhooks = json.loads(env_config)
                except Exception as e_cfg:
                    logger.error(f"自动同步时解析环境变量 TEAM_WEBHOOKS_JSON 失败: {e_cfg}")

            if team_name and team_name in team_webhooks:
                custom_cfg = team_webhooks[team_name]
                if custom_cfg.get("template_id"):
                    template_id = custom_cfg["template_id"]

            if not template_id:
                template_id = settings.DINGTALK_WEEKLY_REPORT_TEMPLATE_ID

            if not template_id:
                logger.error("自动同步周报到钉钉失败: 未配置任何钉钉日志模板ID")
                return

            if template_id == "19cab0d8aa4c349cb1df85146edac9cf":
                template_id = "19eab0d8aa4e349cb1df85146edac9cf"

            # 4. 获取钉钉 ID
            dingtalk_userid = db_user.dingtalk_id
            if not dingtalk_userid:
                dingtalk_userid = await dingtalk_client.get_user_by_mobile(db_user.phone)
                if dingtalk_userid:
                    db_user.dingtalk_id = dingtalk_userid
                    await db.commit()
                else:
                    logger.error(f"自动同步周报到钉钉失败: 用户 {db_user.name} 的手机号 {db_user.phone} 无法在钉钉中匹配到工号")
                    return

            # 5. 映射组装表单内容
            is_marketing = db_user.position_type == "marketing"

            plan_val = report.sales_plan if is_marketing else report.delivery_plan
            actual_val = report.sales_actual if is_marketing else report.delivery_actual
            rate_val = report.sales_rate if is_marketing else report.delivery_rate
            highlights_val = report.sales_highlights if is_marketing else report.delivery_highlights
            blockers_val = report.sales_blockers if is_marketing else report.delivery_blockers
            support_val = report.sales_support if is_marketing else report.delivery_support
            next_plan_val = report.next_sales_plan if is_marketing else report.next_delivery_plan

            start_date_str = report.start_date.strftime('%Y-%m-%d') if hasattr(report.start_date, 'strftime') else str(report.start_date)
            end_date_str = report.end_date.strftime('%Y-%m-%d') if hasattr(report.end_date, 'strftime') else str(report.end_date)

            contents = [
                {"key": "本周目标计划", "value": plan_val or ""},
                {"key": "本周实际完成", "value": actual_val or ""},
                {"key": "达成情况", "value": rate_val or "100%"},
                {"key": "本周亮点", "value": highlights_val or "无"},
                {"key": "本周卡点", "value": blockers_val or "无"},
                {"key": "是否需要上级支持", "value": support_val or "无"},
                {"key": "下周目标", "value": next_plan_val or ""},
                {"key": "周报日期", "value": f"{start_date_str}至{end_date_str}"}
            ]

            # 6. 计算接收人
            to_userids = []
            stmt_admins = select(DbUser.dingtalk_id).where(
                DbUser.role == "admin",
                DbUser.dingtalk_id.isnot(None),
                DbUser.dingtalk_id != ""
            )
            res_admins = await db.execute(stmt_admins)
            admin_ids = res_admins.scalars().all()
            to_userids.extend(admin_ids)

            has_third_bar = db_user.third_class_bar and db_user.third_class_bar.strip() != "" and db_user.third_class_bar.strip().lower() != "all"
            if has_third_bar:
                stmt_teammates = select(DbUser.dingtalk_id).where(
                    DbUser.third_class_bar == db_user.third_class_bar.strip(),
                    DbUser.dingtalk_id.isnot(None),
                    DbUser.dingtalk_id != ""
                )
                res_teammates = await db.execute(stmt_teammates)
                teammate_ids = res_teammates.scalars().all()
                to_userids.extend(teammate_ids)
            elif db_user.team_id:
                stmt_teammates = select(DbUser.dingtalk_id).where(
                    DbUser.team_id == db_user.team_id,
                    DbUser.dingtalk_id.isnot(None),
                    DbUser.dingtalk_id != ""
                )
                res_teammates = await db.execute(stmt_teammates)
                teammate_ids = res_teammates.scalars().all()
                to_userids.extend(teammate_ids)

            to_userids = list(set([uid for uid in to_userids if uid and uid != dingtalk_userid]))

            # 7. 调用钉钉工作日志填报 API 填报数据
            success, err_msg = await dingtalk_client.save_report(template_id, dingtalk_userid, contents, to_userids=to_userids)
            if success:
                logger.info(f"自动同步周报至钉钉工作日志成功: {db_user.name}")
            else:
                logger.error(f"自动同步周报至钉钉工作日志失败: {err_msg}")
        except Exception as e:
            logger.error(f"自动同步周报后台任务发生异常: {e}", exc_info=True)


@router.get("/weekly/summary/export", summary="导出小组成员周复盘汇总表")
async def export_weekly_reports_summary(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """导出周复盘数据大表为 Excel"""
    from io import BytesIO
    import openpyxl
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    from app.models.user import UserRole
    
    if current_user.role not in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value] and current_user.team_id is not None:
        team_id = current_user.team_id

    from app.models.user import User as DbUser
    
    query = select(
        WeeklyReport.id,
        WeeklyReport.user_id,
        WeeklyReport.start_date,
        WeeklyReport.end_date,
        WeeklyReport.delivery_plan,
        WeeklyReport.sales_plan,
        WeeklyReport.delivery_actual,
        WeeklyReport.sales_actual,
        WeeklyReport.delivery_rate,
        WeeklyReport.sales_rate,
        WeeklyReport.delivery_highlights,
        WeeklyReport.sales_highlights,
        WeeklyReport.delivery_blockers,
        WeeklyReport.sales_blockers,
        WeeklyReport.delivery_support,
        WeeklyReport.sales_support,
        WeeklyReport.next_delivery_plan,
        WeeklyReport.next_sales_plan,
        WeeklyReport.status,
        WeeklyReport.submitted_at,
        WeeklyReport.created_at,
        WeeklyReport.updated_at,
        DbUser.name.label("user_name"),
        DbUser.position_type.label("user_position_type"),
        DbUser.role.label("user_role")
    ).join(DbUser, WeeklyReport.user_id == DbUser.id)
    
    if team_id:
        query = query.where(DbUser.team_id == team_id)
        
    if third_class_bar and third_class_bar != "all":
        query = query.where(DbUser.third_class_bar == third_class_bar)
        
    if user_name:
        if "," in user_name or "，" in user_name:
            from sqlalchemy import or_
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                query = query.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            query = query.where(DbUser.name.like(f"%{user_name}%"))
            
    query = query.where(WeeklyReport.start_date == start_date)
    query = query.order_by(DbUser.name.asc())
    
    res = await db.execute(query)
    rows = res.all()
    
    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "周报汇总"
    
    # 写入表头
    headers = [
        "成员姓名",
        "岗位类型",
        "本周目标计划",
        "本周实际完成",
        "达成率",
        "本周亮点",
        "本周卡点",
        "支持协调需求",
        "下周目标计划",
        "状态",
        "提交时间"
    ]
    ws.append(headers)
    
    from app.models.user import PositionType
    for row in rows:
        is_marketing = (
            row.user_position_type == PositionType.MARKETING.value 
            or row.user_role in [UserRole.TARGET_OFFICER.value, "marketing_staff", "tech_marketing"]
        )
        
        # 根据角色选择字段
        weekly_plan = row.sales_plan if is_marketing else row.delivery_plan
        weekly_actual = row.sales_actual if is_marketing else row.delivery_actual
        weekly_rate = row.sales_rate if is_marketing else row.delivery_rate
        weekly_highlights = row.sales_highlights if is_marketing else row.delivery_highlights
        weekly_blockers = row.sales_blockers if is_marketing else row.delivery_blockers
        weekly_support = row.sales_support if is_marketing else row.delivery_support
        weekly_next_plan = row.next_sales_plan if is_marketing else row.next_delivery_plan
        
        status_str = "已提交" if row.status == "submitted" else "草稿"
        submitted_at_str = row.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if row.submitted_at else "—"
        position_type_str = "营销" if is_marketing else "交付"
        
        ws.append([
            row.user_name,
            position_type_str,
            weekly_plan or "—",
            weekly_actual or "—",
            weekly_rate or "—",
            weekly_highlights or "—",
            weekly_blockers or "—",
            weekly_support or "—",
            weekly_next_plan or "—",
            status_str,
            submitted_at_str
        ])
        
    # 格式化与美化 Excel 样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    font_family = "微软雅黑"
    
    # 样式定义
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    content_font = Font(name=font_family, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 设置表头样式
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    # 设置列宽
    col_widths = {
        1: 15,  # 成员姓名
        2: 12,  # 岗位类型
        3: 38,  # 本周目标计划
        4: 38,  # 本周实际完成
        5: 12,  # 达成率
        6: 38,  # 本周亮点
        7: 38,  # 本周卡点
        8: 38,  # 支持协调需求
        9: 38,  # 下周目标计划
        10: 12, # 状态
        11: 22, # 提交时间
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 设置内容样式与动态自适应行高
    center_cols = {1, 2, 5, 10, 11} # 成员姓名、岗位类型、达成率、状态、提交时间居中，其余靠左顶端对齐
    for r_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = content_font
            cell.border = thin_border
            if c_idx in center_cols:
                cell.alignment = align_center
            else:
                cell.alignment = align_left_top
                
            val = cell.value
            if val and val != "—":
                val_str = str(val)
                col_width = col_widths.get(c_idx, 15)
                char_per_line = max(5, int(col_width * 0.5))
                segments = val_str.split("\n")
                lines_count = 0
                for seg in segments:
                    seg_len = 0.0
                    for char in seg:
                        if ord(char) > 127:
                            seg_len += 1.0
                        else:
                            seg_len += 0.5
                    lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
                max_lines = max(max_lines, lines_count)
                
        ws.row_dimensions[r_idx].height = max(24, max_lines * 16 + 10)
        
    # 冻结首行首列 (B2单元格左上角被冻结)
    ws.freeze_panes = "B2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 计算文件名范围前缀
    scope_str = "公司全体"
    if third_class_bar and third_class_bar != "all":
        from app.models.organization import Team as DbTeam
        from app.models.user import User as DbUser
        team_name_str = ""
        if team_id:
            stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        if not team_name_str:
            stmt_team = select(DbTeam.name).join(DbUser, DbUser.team_id == DbTeam.id).where(DbUser.third_class_bar == third_class_bar).limit(1)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        
        if team_name_str:
            scope_str = f"{team_name_str}_{third_class_bar}"
        else:
            scope_str = third_class_bar
    elif team_id:
        from app.models.organization import Team as DbTeam
        stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
        res_team = await db.execute(stmt_team)
        team_name_str = res_team.scalar() or ""
        if team_name_str:
            scope_str = team_name_str
            
    filename = f"{scope_str}_团队周报汇总_{start_date}.xlsx"
    filename_encoded = quote(filename)
    
    # 审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="EXPORT",
        target_module="weekly_report",
        target_id=0,
        description=f"导出了小组成员周复盘汇总表，日期: {start_date}"
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )


@router.get("/weekly/crm-summary/export", summary="导出团队 CRM 业务数据汇总")
async def export_weekly_reports_crm_summary(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """导出团队 CRM 业务数据汇总表为 Excel"""
    from io import BytesIO
    import openpyxl
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    from sqlalchemy import or_, and_
    from app.models.user import User as DbUser, PositionType, UserRole
    from app.models.organization import Team as DbTeam
    from fastapi.concurrency import run_in_threadpool
    import asyncio
    
    # 1. 确定当前用户可看的人员范围
    stmt = select(
        DbUser.id,
        DbUser.name,
        DbUser.third_class_bar,
        DbUser.position_type,
        DbUser.role,
        DbTeam.name.label("team_name")
    ).outerjoin(DbTeam, DbUser.team_id == DbTeam.id).where(DbUser.is_active == True)
    
    if current_user.role in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value]:
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
    else:
        conditions = [DbUser.id == current_user.id]
        if current_user.team_id is not None:
            conditions.append(DbUser.team_id == current_user.team_id)
        if current_user.third_class_bar:
            conditions.append(and_(
                DbUser.third_class_bar == current_user.third_class_bar,
                DbUser.third_class_bar != None,
                DbUser.third_class_bar != ""
            ))
        stmt = stmt.where(or_(*conditions))
        
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
            
    if third_class_bar and third_class_bar != "all":
        stmt = stmt.where(DbUser.third_class_bar == third_class_bar)
        
    if user_name:
        if "," in user_name or "，" in user_name:
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                stmt = stmt.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            stmt = stmt.where(DbUser.name.like(f"%{user_name}%"))
            
    stmt = stmt.order_by(DbUser.name.asc())
    res = await db.execute(stmt)
    rows = res.all()
    
    # 2. 并发拉取 CRM 数据
    async def fetch_user_crm_data(user_row):
        is_marketing = (
            user_row.position_type == PositionType.MARKETING 
            or user_row.role in [UserRole.TARGET_OFFICER, UserRole.MARKETING_STAFF, UserRole.TECH_MARKETING]
        )
        # 调用跨库查询函数
        crm_data = await run_in_threadpool(
            sync_extract_crm_data, 
            user_row.name,
            start_date, 
            is_marketing
        )
        return {
            "user_name": user_row.name,
            "third_class_bar": user_row.third_class_bar,
            "team_name": user_row.team_name,
            "crm_active_projects": crm_data.get("crm_active_projects"),
            "crm_milestone_tasks": crm_data.get("crm_milestone_tasks"),
            "crm_suspended_projects": crm_data.get("crm_suspended_projects"),
            "crm_no_contract_warning": crm_data.get("crm_no_contract_warning"),
            "crm_unbilled_warning": crm_data.get("crm_unbilled_warning"),
            "crm_unreceived_warning": crm_data.get("crm_unreceived_warning"),
            "crm_health_diagnosis": crm_data.get("crm_health_diagnosis")
        }
        
    tasks = [fetch_user_crm_data(row) for row in rows]
    items = await asyncio.gather(*tasks) if tasks else []
    
    # 3. 创建 Excel 并写入数据
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "团队CRM信息"
    
    headers = [
        "成员姓名",
        "三级巴",
        "归属战队",
        "正在实施项目进度",
        "里程碑与交付动作明细",
        "暂停或异常挂起项目",
        "预设立立警 (超期未签合同)",
        "交付卡点 (有进度未开票)",
        "收欠款预警 (已开票未回款)",
        "饱和度与健康度诊断"
    ]
    ws.append(headers)
    
    for item in items:
        ws.append([
            item["user_name"],
            item["third_class_bar"] or "—",
            item["team_name"] or "—",
            item["crm_active_projects"] or "—",
            item["crm_milestone_tasks"] or "—",
            item["crm_suspended_projects"] or "—",
            item["crm_no_contract_warning"] or "—",
            item["crm_unbilled_warning"] or "—",
            item["crm_unreceived_warning"] or "—",
            item["crm_health_diagnosis"] or "—"
        ])
        
    # 格式化与美化 Excel 样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    font_family = "微软雅黑"
    
    # 样式定义
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    content_font = Font(name=font_family, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 设置表头样式
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    # 设置列宽
    col_widths = {
        1: 15,  # 成员姓名
        2: 15,  # 三级巴
        3: 15,  # 归属战队
        4: 40,  # 正在实施项目进度
        5: 40,  # 里里程碑与交付动作明细
        6: 40,  # 暂停或异常挂起项目
        7: 40,  # 预设立立警 (超期未签合同)
        8: 40,  # 交付卡点 (有进度未开票)
        9: 40,  # 收欠款预警 (已开票未回款)
        10: 40, # 饱和度与健康度诊断
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 设置内容样式与动态自适应行高
    center_cols = {1, 2, 3} # 成员姓名、三级巴、归属战队居中，其余靠左顶端对齐
    for r_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = content_font
            cell.border = thin_border
            if c_idx in center_cols:
                cell.alignment = align_center
            else:
                cell.alignment = align_left_top
                
            val = cell.value
            if val and val != "—":
                val_str = str(val)
                col_width = col_widths.get(c_idx, 15)
                char_per_line = max(5, int(col_width * 0.5))
                segments = val_str.split("\n")
                lines_count = 0
                for seg in segments:
                    seg_len = 0.0
                    for char in seg:
                        if ord(char) > 127:
                            seg_len += 1.0
                        else:
                            seg_len += 0.5
                    lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
                max_lines = max(max_lines, lines_count)
                
        ws.row_dimensions[r_idx].height = max(24, max_lines * 16 + 10)
        
    # 冻结首行首列 (B2单元格左上角被冻结)
    ws.freeze_panes = "B2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 计算文件名范围前缀
    scope_str = "公司全体"
    if third_class_bar and third_class_bar != "all":
        from app.models.organization import Team as DbTeam
        from app.models.user import User as DbUser
        team_name_str = ""
        if team_id:
            stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        if not team_name_str:
            stmt_team = select(DbTeam.name).join(DbUser, DbUser.team_id == DbTeam.id).where(DbUser.third_class_bar == third_class_bar).limit(1)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        
        if team_name_str:
            scope_str = f"{team_name_str}_{third_class_bar}"
        else:
            scope_str = third_class_bar
    elif team_id:
        from app.models.organization import Team as DbTeam
        stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
        res_team = await db.execute(stmt_team)
        team_name_str = res_team.scalar() or ""
        if team_name_str:
            scope_str = team_name_str
            
    filename = f"{scope_str}_团队CRM汇总_{start_date}.xlsx"
    filename_encoded = quote(filename)
    
    # 审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="EXPORT",
        target_module="weekly_crm_report",
        target_id=0,
        description=f"导出了团队 CRM 业务数据汇总表，日期: {start_date}"
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )


@router.get("/weekly/summary/export-horizontal", summary="横版导出小组成员周复盘汇总表")
async def export_weekly_reports_summary_horizontal(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """横版导出周复盘数据大表为 Excel"""
    from io import BytesIO
    import openpyxl
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    from app.models.user import UserRole
    
    if current_user.role not in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value] and current_user.team_id is not None:
        team_id = current_user.team_id

    from app.models.user import User as DbUser
    
    query = select(
        WeeklyReport.id,
        WeeklyReport.user_id,
        WeeklyReport.start_date,
        WeeklyReport.end_date,
        WeeklyReport.delivery_plan,
        WeeklyReport.sales_plan,
        WeeklyReport.delivery_actual,
        WeeklyReport.sales_actual,
        WeeklyReport.delivery_rate,
        WeeklyReport.sales_rate,
        WeeklyReport.delivery_highlights,
        WeeklyReport.sales_highlights,
        WeeklyReport.delivery_blockers,
        WeeklyReport.sales_blockers,
        WeeklyReport.delivery_support,
        WeeklyReport.sales_support,
        WeeklyReport.next_delivery_plan,
        WeeklyReport.next_sales_plan,
        WeeklyReport.status,
        WeeklyReport.submitted_at,
        WeeklyReport.created_at,
        WeeklyReport.updated_at,
        DbUser.name.label("user_name"),
        DbUser.position_type.label("user_position_type"),
        DbUser.role.label("user_role")
    ).join(DbUser, WeeklyReport.user_id == DbUser.id)
    
    if team_id:
        query = query.where(DbUser.team_id == team_id)
        
    if third_class_bar and third_class_bar != "all":
        query = query.where(DbUser.third_class_bar == third_class_bar)
        
    if user_name:
        if "," in user_name or "，" in user_name:
            from sqlalchemy import or_
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                query = query.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            query = query.where(DbUser.name.like(f"%{user_name}%"))
            
    query = query.where(WeeklyReport.start_date == start_date)
    query = query.order_by(DbUser.name.asc())
    
    res = await db.execute(query)
    rows = res.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "周报汇总横版"
    
    # 构造表头：首列是维度，后面每一列是一个人的名字
    headers = ["指标/维度"]
    from app.models.user import PositionType
    for row in rows:
        is_marketing = (
            row.user_position_type == PositionType.MARKETING.value 
            or row.user_role in [UserRole.TARGET_OFFICER.value, "marketing_staff", "tech_marketing"]
        )
        post_str = "营销岗" if is_marketing else "交付岗"
        headers.append(f"{row.user_name} ({post_str})")
    
    ws.append(headers)
    
    # 准备 6 行对应 6 个维度
    dimensions = [
        ("🎯 本周目标计划", "plan"),
        ("🔥 本周实际完成", "actual"),
        ("🏆 本周工作亮点", "highlights"),
        ("🚧 本周工作卡点/难点", "blockers"),
        ("🤝 需要支持协调", "support"),
        ("🚀 下周工作目标", "next_plan")
    ]
    
    for label, key in dimensions:
        row_cells = [label]
        for row in rows:
            is_marketing = (
                row.user_position_type == PositionType.MARKETING.value 
                or row.user_role in [UserRole.TARGET_OFFICER.value, "marketing_staff", "tech_marketing"]
            )
            val = ""
            if key == "plan":
                val = row.sales_plan if is_marketing else row.delivery_plan
            elif key == "actual":
                val = row.sales_actual if is_marketing else row.delivery_actual
            elif key == "highlights":
                val = row.sales_highlights if is_marketing else row.delivery_highlights
            elif key == "blockers":
                val = row.sales_blockers if is_marketing else row.delivery_blockers
            elif key == "support":
                val = row.sales_support if is_marketing else row.delivery_support
            elif key == "next_plan":
                val = row.next_sales_plan if is_marketing else row.next_delivery_plan
            
            row_cells.append(val or "—")
        ws.append(row_cells)
        
    # 格式化与美化 Excel 样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    font_family = "微软雅黑"
    
    # 样式定义
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    content_font = Font(name=font_family, size=10)
    dim_font = Font(name=font_family, size=10, bold=True, color="333333")
    dim_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    dim_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 设置表头样式
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    # 设置列宽
    ws.column_dimensions[get_column_letter(1)].width = 24  # 指标/维度
    for c_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 45  # 人员内容列

    # 设置内容样式与动态自适应行高
    for r_idx in range(2, ws.max_row + 1):
        max_lines = 1
        
        # 第一列：维度标签
        cell_dim = ws.cell(row=r_idx, column=1)
        cell_dim.font = dim_font
        cell_dim.fill = dim_fill
        cell_dim.alignment = dim_align
        cell_dim.border = thin_border
        
        val_dim = cell_dim.value
        if val_dim and val_dim != "—":
            val_str = str(val_dim)
            char_per_line = max(5, int(24 * 0.5))
            segments = val_str.split("\n")
            lines_count = 0
            for seg in segments:
                seg_len = 0.0
                for char in seg:
                    if ord(char) > 127:
                        seg_len += 1.0
                    else:
                        seg_len += 0.5
                lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
            max_lines = max(max_lines, lines_count)
        
        # 后面每一列：成员的周报维度具体内容
        for c_idx in range(2, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = content_font
            cell.alignment = align_left_top
            cell.border = thin_border
            
            val = cell.value
            if val and val != "—":
                val_str = str(val)
                char_per_line = max(5, int(45 * 0.5))
                segments = val_str.split("\n")
                lines_count = 0
                for seg in segments:
                    seg_len = 0.0
                    for char in seg:
                        if ord(char) > 127:
                            seg_len += 1.0
                        else:
                            seg_len += 0.5
                    lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
                max_lines = max(max_lines, lines_count)
                
        ws.row_dimensions[r_idx].height = max(24, max_lines * 16 + 10)
        
    # 冻结首行首列 (B2单元格左上角被冻结)
    ws.freeze_panes = "B2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 计算文件名范围前缀
    scope_str = "公司全体"
    if third_class_bar and third_class_bar != "all":
        from app.models.organization import Team as DbTeam
        from app.models.user import User as DbUser
        team_name_str = ""
        if team_id:
            stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        if not team_name_str:
            stmt_team = select(DbTeam.name).join(DbUser, DbUser.team_id == DbTeam.id).where(DbUser.third_class_bar == third_class_bar).limit(1)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        
        if team_name_str:
            scope_str = f"{team_name_str}_{third_class_bar}"
        else:
            scope_str = third_class_bar
    elif team_id:
        from app.models.organization import Team as DbTeam
        stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
        res_team = await db.execute(stmt_team)
        team_name_str = res_team.scalar() or ""
        if team_name_str:
            scope_str = team_name_str
            
    filename = f"{scope_str}_团队周报汇总_横版_{start_date}.xlsx"
    filename_encoded = quote(filename)
    
    await log_action(
        db=db,
        user=current_user,
        action_type="EXPORT",
        target_module="weekly_report",
        target_id=0,
        description=f"横版导出了小组成员周报，日期: {start_date}"
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )


@router.get("/weekly/crm-summary/export-horizontal", summary="横版导出团队 CRM 业务数据汇总")
async def export_weekly_reports_crm_summary_horizontal(
    start_date: date = Query(..., description="周开始日期(周一)"),
    team_id: int | None = Query(None, description="按战队/小组筛选"),
    third_class_bar: str | None = Query(None, description="按三级巴筛选"),
    user_name: str | None = Query(None, description="按人员姓名筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("view_weekly_reports")),
):
    """横版导出团队 CRM 业务数据汇总表为 Excel"""
    from io import BytesIO
    import openpyxl
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    from sqlalchemy import or_, and_
    from app.models.user import User as DbUser, PositionType, UserRole
    from app.models.organization import Team as DbTeam
    from fastapi.concurrency import run_in_threadpool
    import asyncio
    
    stmt = select(
        DbUser.id,
        DbUser.name,
        DbUser.third_class_bar,
        DbUser.position_type,
        DbUser.role,
        DbTeam.name.label("team_name")
    ).outerjoin(DbTeam, DbUser.team_id == DbTeam.id).where(DbUser.is_active == True)
    
    if current_user.role in [UserRole.ADMIN.value, UserRole.TARGET_OFFICER.value]:
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
    else:
        conditions = [DbUser.id == current_user.id]
        if current_user.team_id is not None:
            conditions.append(DbUser.team_id == current_user.team_id)
        if current_user.third_class_bar:
            conditions.append(and_(
                DbUser.third_class_bar == current_user.third_class_bar,
                DbUser.third_class_bar != None,
                DbUser.third_class_bar != ""
            ))
        stmt = stmt.where(or_(*conditions))
        
        if team_id:
            stmt = stmt.where(DbUser.team_id == team_id)
            
    if third_class_bar and third_class_bar != "all":
        stmt = stmt.where(DbUser.third_class_bar == third_class_bar)
        
    if user_name:
        if "," in user_name or "，" in user_name:
            names = [n.strip() for n in user_name.replace("，", ",").split(",") if n.strip()]
            if names:
                stmt = stmt.where(or_(*[DbUser.name.like(f"%{name}%") for name in names]))
        else:
            stmt = stmt.where(DbUser.name.like(f"%{user_name}%"))
            
    stmt = stmt.order_by(DbUser.name.asc())
    res = await db.execute(stmt)
    rows = res.all()
    
    async def fetch_user_crm_data(user_row):
        is_marketing = (
            user_row.position_type == PositionType.MARKETING 
            or user_row.role in [UserRole.TARGET_OFFICER, UserRole.MARKETING_STAFF, UserRole.TECH_MARKETING]
        )
        crm_data = await run_in_threadpool(
            sync_extract_crm_data, 
            user_row.name,
            start_date, 
            is_marketing
        )
        return {
            "user_name": user_row.name,
            "crm_active_projects": crm_data.get("crm_active_projects"),
            "crm_milestone_tasks": crm_data.get("crm_milestone_tasks"),
            "crm_suspended_projects": crm_data.get("crm_suspended_projects"),
            "crm_no_contract_warning": crm_data.get("crm_no_contract_warning"),
            "crm_unbilled_warning": crm_data.get("crm_unbilled_warning"),
            "crm_unreceived_warning": crm_data.get("crm_unreceived_warning"),
            "crm_health_diagnosis": crm_data.get("crm_health_diagnosis")
        }
        
    tasks = [fetch_user_crm_data(row) for row in rows]
    items = await asyncio.gather(*tasks) if tasks else []
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM汇总横版"
    
    # 构造表头：列是人名
    headers = ["指标/维度"]
    for item in items:
        headers.append(item["user_name"])
    ws.append(headers)
    
    dimensions = [
        ("💻 正在实施项目进度", "crm_active_projects"),
        ("🎯 里程碑与交付动作", "crm_milestone_tasks"),
        ("⚠️ 暂停或异常挂起项目", "crm_suspended_projects"),
        ("🔴 合同超期未签预警", "crm_no_contract_warning"),
        ("🟡 有进度未开票卡点", "crm_unbilled_warning"),
        ("🔴 已开票未回款预警", "crm_unreceived_warning"),
        ("🩺 饱和度与健康度诊断", "crm_health_diagnosis")
    ]
    
    for label, key in dimensions:
        row_cells = [label]
        for item in items:
            row_cells.append(item[key] or "—")
        ws.append(row_cells)
        
    # 格式化与美化 Excel 样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    font_family = "微软雅黑"
    
    # 样式定义
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    content_font = Font(name=font_family, size=10)
    dim_font = Font(name=font_family, size=10, bold=True, color="333333")
    dim_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    dim_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 设置表头样式
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    # 设置列宽
    ws.column_dimensions[get_column_letter(1)].width = 24  # 指标/维度
    for c_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 45  # 人员内容列

    # 设置内容样式与动态自适应行高
    for r_idx in range(2, ws.max_row + 1):
        max_lines = 1
        
        # 第一列：维度标签
        cell_dim = ws.cell(row=r_idx, column=1)
        cell_dim.font = dim_font
        cell_dim.fill = dim_fill
        cell_dim.alignment = dim_align
        cell_dim.border = thin_border
        
        val_dim = cell_dim.value
        if val_dim and val_dim != "—":
            val_str = str(val_dim)
            char_per_line = max(5, int(24 * 0.5))
            segments = val_str.split("\n")
            lines_count = 0
            for seg in segments:
                seg_len = 0.0
                for char in seg:
                    if ord(char) > 127:
                        seg_len += 1.0
                    else:
                        seg_len += 0.5
                lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
            max_lines = max(max_lines, lines_count)
        
        # 后面每一列：成员的 CRM 维度具体内容
        for c_idx in range(2, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = content_font
            cell.alignment = align_left_top
            cell.border = thin_border
            
            val = cell.value
            if val and val != "—":
                val_str = str(val)
                char_per_line = max(5, int(45 * 0.5))
                segments = val_str.split("\n")
                lines_count = 0
                for seg in segments:
                    seg_len = 0.0
                    for char in seg:
                        if ord(char) > 127:
                            seg_len += 1.0
                        else:
                            seg_len += 0.5
                    lines_count += max(1, int(seg_len / char_per_line) + (1 if seg_len % char_per_line > 0 else 0))
                max_lines = max(max_lines, lines_count)
                
        ws.row_dimensions[r_idx].height = max(24, max_lines * 16 + 10)
        
    # 冻结首行首列 (B2单元格左上角被冻结)
    ws.freeze_panes = "B2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 计算文件名范围前缀
    scope_str = "公司全体"
    if third_class_bar and third_class_bar != "all":
        from app.models.organization import Team as DbTeam
        from app.models.user import User as DbUser
        team_name_str = ""
        if team_id:
            stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        if not team_name_str:
            stmt_team = select(DbTeam.name).join(DbUser, DbUser.team_id == DbTeam.id).where(DbUser.third_class_bar == third_class_bar).limit(1)
            res_team = await db.execute(stmt_team)
            team_name_str = res_team.scalar() or ""
        
        if team_name_str:
            scope_str = f"{team_name_str}_{third_class_bar}"
        else:
            scope_str = third_class_bar
    elif team_id:
        from app.models.organization import Team as DbTeam
        stmt_team = select(DbTeam.name).where(DbTeam.id == team_id)
        res_team = await db.execute(stmt_team)
        team_name_str = res_team.scalar() or ""
        if team_name_str:
            scope_str = team_name_str
            
    filename = f"{scope_str}_团队CRM汇总_横版_{start_date}.xlsx"
    filename_encoded = quote(filename)
    
    await log_action(
        db=db,
        user=current_user,
        action_type="EXPORT",
        target_module="weekly_crm_report",
        target_id=0,
        description=f"横版导出了团队 CRM 业务数据，日期: {start_date}"
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )
