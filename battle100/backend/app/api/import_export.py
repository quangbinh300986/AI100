"""
Excel导入导出接口
提供数据批量导入导出功能
"""

from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
import pymysql
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.models.user import User, UserRole
from app.models.goal import PersonalGoal, TeamGoal, WeeklyTarget, GoalType
from app.models.organization import Team
from fastapi import Request
from app.api.deps import get_current_user, require_roles, require_permission
from app.services.audit_service import log_action, to_dict

# 动态权限代理拦截：依据请求路径，将导入导出动作智能分流到 settings (用户/钉钉操作) 或 goals (周目标导入/导出) 上
def dynamic_require_roles(*roles):
    async def import_export_permission_dependency(
        request: Request,
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
    ) -> User:
        if current_user.role == UserRole.ADMIN:
            return current_user
            
        path = request.url.path
        
        # 1. 用户导入导出及钉钉同步映射为 settings 相关的管理权限
        if "users" in path:
            perm = "manage_user_roles"
        # 2. 周目标 Excel 导入映射为 import_weekly_targets
        elif "goals" in path and "import" in path:
            perm = "import_weekly_targets"
        # 3. 其它（如周目标导出）映射为 view_goals 即可
        else:
            perm = "view_goals"
            
        checker = require_permission(perm)
        return await checker(current_user, db)
        
    return import_export_permission_dependency

require_roles = dynamic_require_roles

router = APIRouter(prefix="/import-export", tags=["导入导出"])


@router.post("/users/import", summary="批量导入用户")
async def import_users(
    file: UploadFile = File(..., description="Excel文件"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
):
    """
    从Excel文件批量导入用户
    Excel格式：姓名 | 手机号 | 岗位 | 岗位类型 | 角色
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件")

    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    imported_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0] or not row[1]:
                continue  # 跳过空行

            name = str(row[0]).strip()
            phone = str(row[1]).strip()

            # 检查手机号是否已存在
            existing = await db.execute(
                select(User).where(User.phone == phone)
            )
            if existing.scalar_one_or_none():
                errors.append(f"第{row_idx}行: 手机号{phone}已存在")
                continue

            from app.services.auth_service import hash_password
            user = User(
                name=name,
                phone=phone,
                password_hash=hash_password("123456"),  # 默认密码
                position=str(row[2]).strip() if row[2] else None,
                position_type=str(row[3]).strip() if row[3] else None,
                role=str(row[4]).strip() if row[4] else "staff",
                third_class_bar=str(row[5]).strip() if len(row) > 5 and row[5] else None,
            )
            db.add(user)
            imported_count += 1
        except Exception as e:
            errors.append(f"第{row_idx}行导入失败: {str(e)}")

    await db.flush()

    # 记录审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="IMPORT",
        target_module="user",
        target_id="0",
        description=f"Excel批量导入用户，成功导入 {imported_count} 个用户",
        before_state=None,
        after_state={"imported_count": imported_count, "errors": errors},
    )

    return {
        "message": f"成功导入{imported_count}个用户",
        "imported_count": imported_count,
        "errors": errors,
    }


@router.get("/users/export", summary="导出用户列表")
async def export_users(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
):
    """将用户列表导出为Excel文件"""
    result = await db.execute(select(User).order_by(User.id))
    users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"

    # 表头
    headers = ["ID", "姓名", "手机号", "岗位", "岗位类型", "角色", "三级巴", "战队ID", "是否激活", "创建时间"]
    ws.append(headers)

    # 数据行
    for user in users:
        ws.append([
            user.id,
            user.name,
            user.phone,
            user.position,
            user.position_type,
            user.role,
            user.third_class_bar,
            user.team_id,
            "是" if user.is_active else "否",
            user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
        ])

    # 返回文件流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_export.xlsx"},
    )


import re
from datetime import date, datetime

def parse_week_date_range(text: str) -> tuple[date, date] | None:
    if not text:
        return None
    text = str(text).strip()
    # 匹配类似于 "6月1日-6月7日" 或 "5月29日-6月5日"
    pattern = r"(\d+)\s*月\s*(\d+)\s*日\s*-\s*(\d+)\s*月\s*(\d+)\s*日"
    match = re.search(pattern, text)
    if match:
        m1, d1, m2, d2 = map(int, match.groups())
        return date(2026, m1, d1), date(2026, m2, d2)
    
    # 兼容没有"日"的格式，如 "6月1-6月7"
    pattern2 = r"(\d+)\s*月\s*(\d+)\s*-\s*(\d+)\s*月\s*(\d+)"
    match2 = re.search(pattern2, text)
    if match2:
        m1, d1, m2, d2 = map(int, match2.groups())
        return date(2026, m1, d1), date(2026, m2, d2)
        
    return None

@router.post("/goals/weekly/import", summary="批量导入周分解目标")
async def import_goals_weekly(
    file: UploadFile = File(..., description="周分解目标Excel文件"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.TARGET_OFFICER)),
):
    """
    导入各战队每周的基础/挑战目标（附件1）
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件")

    content = await file.read()
    # 启用 data_only=True，使得公式直接返回计算后的数值结果
    wb = load_workbook(BytesIO(content), data_only=True)
    
    # 优先显式查找"新签目标"工作表，以防活跃Sheet不是该目标页
    if "新签目标" in wb.sheetnames:
        ws = wb["新签目标"]
    else:
        ws = wb.active

    # 获取战队映射缓存
    team_res = await db.execute(select(Team))
    all_teams = team_res.scalars().all()
    team_by_name = {t.name: t.id for t in all_teams}

    # 1. 扫描表头，建立每一列对应的周次与起止日期
    week_columns = {}
    week_counter = 1
    
    header_found = False
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                range_res = parse_week_date_range(str(val))
                if range_res:
                    start_date, end_date = range_res
                    if c not in week_columns:
                        week_columns[c] = {
                            "week_number": week_counter,
                            "week_start": start_date,
                            "week_end": end_date
                        }
                        week_counter += 1
                        header_found = True

    if not header_found or not week_columns:
        raise HTTPException(status_code=400, detail="未能识别Excel中的周次日期列，请确认是否有类似于“6月1日-6月7日”的表头")

    imported_count = 0
    errors = []
    
    # 2. 遍历每一行，读取战队和指标数据
    current_mode = "base"
    last_team_name = ""

    for r in range(1, ws.max_row + 1):
        # 扫描前4列的数据来确定当前的模式切换
        row_values = [ws.cell(row=r, column=c).value for c in range(1, 5)]
        row_str = " ".join([str(v) for v in row_values if v]).lower()
        
        if "挑战" in row_str or "高目标" in row_str:
            current_mode = "challenge"
            continue
        elif "基础" in row_str or "保底" in row_str or "奋斗目标" in row_str:
            current_mode = "base"
            continue

        # 检查是否为汇总/总计行，防止数据污染
        zone_val = ws.cell(row=r, column=2).value
        zone_str = str(zone_val).strip() if zone_val else ""
        if zone_str and any(x in zone_str for x in ["中地顾问", "合计", "总计"]):
            last_team_name = ""
            continue

        # 对应真实表格结构：
        # B列(第2列)是战区；C列(第3列)是战队名；D列(第4列)是指标名
        team_cell_val = ws.cell(row=r, column=3).value
        indicator_cell_val = ws.cell(row=r, column=4).value

        if team_cell_val:
            last_team_name = str(team_cell_val).strip()
            
        if not last_team_name or not indicator_cell_val:
            continue

        team_name = last_team_name
        matched_team_id = None
        for name, tid in team_by_name.items():
            if name in team_name or team_name in name:
                matched_team_id = tid
                break
                
        if not matched_team_id:
            continue

        indicator = str(indicator_cell_val).strip()
        is_marketing = "营销" in indicator
        is_delivery = "交付" in indicator
        
        if not is_marketing and not is_delivery:
            continue

        for col_idx, info in week_columns.items():
            cell_val = ws.cell(row=r, column=col_idx).value
            target_val = 0.0
            if cell_val is not None:
                try:
                    val_str = str(cell_val).replace("万", "").replace("次", "").strip()
                    target_val = float(val_str)
                except ValueError:
                    target_val = 0.0

            week_num = info["week_number"]
            wt_res = await db.execute(
                select(WeeklyTarget).where(
                    WeeklyTarget.team_id == matched_team_id,
                    WeeklyTarget.week_number == week_num
                )
            )
            wt = wt_res.scalar_one_or_none()
            
            if not wt:
                wt = WeeklyTarget(
                    team_id=matched_team_id,
                    week_number=week_num,
                    week_start=info["week_start"],
                    week_end=info["week_end"]
                )
                db.add(wt)
                await db.flush()

            if current_mode == "base":
                if is_marketing:
                    wt.marketing_base_target = target_val
                else:
                    wt.delivery_base_target = target_val
            else:
                if is_marketing:
                    wt.marketing_challenge_target = target_val
                else:
                    wt.delivery_challenge_target = target_val
            
            imported_count += 1

    await db.flush()

    # 3. 自动累加聚合更新战队总目标 (TeamGoal)
    try:
        from sqlalchemy import func
        from app.models.goal import TeamGoalCategory
        
        # 获取所有战队
        all_teams_res = await db.execute(select(Team))
        teams = all_teams_res.scalars().all()
        for t in teams:
            # 聚合营销保底和挑战
            marketing_base = await db.scalar(
                select(func.coalesce(func.sum(WeeklyTarget.marketing_base_target), 0))
                .where(WeeklyTarget.team_id == t.id)
            ) or 0.0
            marketing_challenge = await db.scalar(
                select(func.coalesce(func.sum(WeeklyTarget.marketing_challenge_target), 0))
                .where(WeeklyTarget.team_id == t.id)
            ) or 0.0
            
            # 聚合交付保底和挑战
            delivery_base = await db.scalar(
                select(func.coalesce(func.sum(WeeklyTarget.delivery_base_target), 0))
                .where(WeeklyTarget.team_id == t.id)
            ) or 0.0
            delivery_challenge = await db.scalar(
                select(func.coalesce(func.sum(WeeklyTarget.delivery_challenge_target), 0))
                .where(WeeklyTarget.team_id == t.id)
            ) or 0.0
            
            # 更新营销 TeamGoal
            g_m_res = await db.execute(
                select(TeamGoal).where(
                    TeamGoal.team_id == t.id,
                    TeamGoal.category == TeamGoalCategory.MARKETING
                )
            )
            g_m = g_m_res.scalar_one_or_none()
            if not g_m:
                g_m = TeamGoal(team_id=t.id, category=TeamGoalCategory.MARKETING)
            g_m.base_target = marketing_base
            g_m.red_line_target = marketing_challenge
            g_m.gap = max(0.0, marketing_challenge - marketing_base)
            db.add(g_m)
            
            # 更新交付 TeamGoal
            g_d_res = await db.execute(
                select(TeamGoal).where(
                    TeamGoal.team_id == t.id,
                    TeamGoal.category == TeamGoalCategory.DELIVERY
                )
            )
            g_d = g_d_res.scalar_one_or_none()
            if not g_d:
                g_d = TeamGoal(team_id=t.id, category=TeamGoalCategory.DELIVERY)
            g_d.base_target = delivery_base
            g_d.red_line_target = delivery_challenge
            g_d.gap = max(0.0, delivery_challenge - delivery_base)
            db.add(g_d)
            
        await db.flush()
    except Exception as aggregate_err:
        errors.append(f"周分解目标导入成功，但在自动累加更新战队总目标时失败: {str(aggregate_err)}")

    # 记录审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="IMPORT",
        target_module="goal",
        target_id="0",
        description=f"Excel批量导入周分解目标，成功导入 {imported_count} 条数据",
        before_state=None,
        after_state={"imported_count": imported_count, "errors": errors},
    )

    return {
        "message": f"成功导入 {imported_count} 条周度目标分配数据",
        "imported_count": imported_count,
        "errors": errors,
    }


@router.post("/goals/personal/import", summary="批量导入个人与战队多Sheet目标")
async def import_goals_personal(
    file: UploadFile = File(..., description="个人目标Excel文件"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.TARGET_OFFICER)),
):
    """
    读取并导入目标官、营销岗、技术岗等个人多维度指标（附件2, 3, 4）
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)

    imported_count = 0
    errors = []

    TARGET_SHEETS = ["目标官个人目标", "营销岗个人目标", "技术岗个人目标"]
    
    team_res = await db.execute(select(Team))
    all_teams = team_res.scalars().all()
    team_by_name = {t.name: t.id for t in all_teams}

    for sheet_name in wb.sheetnames:
        if not any(x in sheet_name for x in TARGET_SHEETS):
            continue
            
        ws = wb[sheet_name]
        
        columns_mapping = {}
        name_col_idx = 1
        name_col_found = False
        last_major_header = ""
        
        header_rows = []
        for r in range(1, 4):
            header_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
            
        for col_idx in range(ws.max_column):
            val_r1 = header_rows[0][col_idx]
            val_r2 = header_rows[1][col_idx]
            val_r3 = header_rows[2][col_idx]
            
            cell_strs = [str(v).strip() for v in [val_r1, val_r2, val_r3] if v]
            
            # 遇到目标达成明细记录或周统计辅助表格标识，直接终止扫描，拦截右侧说明区
            if any(x in s for s in cell_strs for x in ["目标达成明细", "明细记录", "周统计"]):
                break
                
            if any("姓名" in s for s in cell_strs):
                if not name_col_found:
                    name_col_idx = col_idx + 1
                    name_col_found = True
                continue
                
            major_header = ""
            if val_r1:
                major_header = str(val_r1).strip()
                last_major_header = major_header
            elif val_r2 and not any(x in str(val_r2) for x in ["基础", "挑战", "保底"]):
                major_header = str(val_r2).strip()
                last_major_header = major_header
            else:
                major_header = last_major_header
                
            sub_header = ""
            if val_r2 and any(x in str(val_r2) for x in ["基础", "挑战", "保底"]):
                sub_header = str(val_r2).strip()
            elif val_r3 and any(x in str(val_r3) for x in ["基础", "挑战", "保底"]):
                sub_header = str(val_r3).strip()
                
            goal_type = None
            major_header_lower = major_header.lower()
            
            if any(x in major_header_lower for x in ["合同额", "签约金额", "新签", "续签"]):
                goal_type = GoalType.CONTRACT_AMOUNT
            elif "线索转化率" in major_header_lower:
                goal_type = GoalType.LEADS_CONVERSION_RATE
            elif any(x in major_header_lower for x in ["客户幸福动作", "幸福行动"]):
                goal_type = GoalType.HAPPINESS_ACTION
            elif "新客户" in major_header_lower:
                goal_type = GoalType.NEW_CUSTOMER_COUNT
            elif "幸福故事" in major_header_lower:
                goal_type = GoalType.HAPPINESS_STORY_COUNT
            elif any(x in major_header_lower for x in ["铁三角", "协同"]):
                goal_type = GoalType.TRIANGLE_COUNT
            elif "有效线索" in major_header_lower:
                goal_type = GoalType.LEADS_COUNT
                
            target_type = None
            if any(x in sub_header for x in ["基础", "保底"]):
                target_type = "base"
            elif "挑战" in sub_header:
                target_type = "challenge"
                
            if goal_type and target_type:
                columns_mapping[col_idx + 1] = {
                    "goal_type": goal_type,
                    "target_type": target_type
                }

        for r in range(4, ws.max_row + 1):
            name_val = ws.cell(row=r, column=name_col_idx).value
            if not name_val:
                continue
                
            name = str(name_val).strip()
            if any(x in name for x in ["合计", "平均", "目标", "总计"]):
                continue

            user_res = await db.execute(select(User).where(User.name == name))
            user = user_res.scalar_one_or_none()
            if not user:
                errors.append(f"工作表 '{sheet_name}' 第 {r} 行: 未在系统中查到员工【{name}】，跳过该行个人目标设定")
                continue

            for col_idx, mapping in columns_mapping.items():
                cell_val = ws.cell(row=r, column=col_idx).value
                if cell_val is None:
                    continue
                    
                target_val = None
                val_str = str(cell_val).replace("万", "").replace("次", "").replace("%", "").strip()
                if val_str:
                    val_clean = "".join([c for c in val_str if c.isdigit() or c == "."])
                    if val_clean:
                        try:
                            target_val = float(val_clean)
                        except ValueError:
                            pass
                    
                    if target_val is None:
                        # 拥有文本但无法解析出数字，例如“按照清远战队要求”
                        errors.append(f"工作表 '{sheet_name}' 第 {r} 行: 员工【{name}】的指标值为 '{cell_val}'，因无法解析为有效数字被跳过转入")
                        continue
                else:
                    # 单元格为空白字符
                    continue

                goal_type = mapping["goal_type"]
                target_type = mapping["target_type"]

                pg_res = await db.execute(
                    select(PersonalGoal).where(
                        PersonalGoal.user_id == user.id,
                        PersonalGoal.goal_type == goal_type
                    )
                )
                pg = pg_res.scalar_one_or_none()
                
                if not pg:
                    pg = PersonalGoal(
                        user_id=user.id,
                        goal_type=goal_type
                    )
                    db.add(pg)
                    await db.flush()

                if target_type == "base":
                    pg.base_target = target_val
                else:
                    pg.challenge_target = target_val
                
                imported_count += 1

    await db.flush()

    # 记录审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="IMPORT",
        target_module="goal",
        target_id="0",
        description=f"Excel批量导入个人与战队多Sheet目标，成功同步 {imported_count} 条数据",
        before_state=None,
        after_state={"imported_count": imported_count, "errors": errors},
    )

    return {
        "message": f"成功同步个人与战队多 Sheet 个人目标 {imported_count} 条",
        "imported_count": imported_count,
        "errors": errors,
    }


# ===== 钉钉通讯录直连同步接口 =====
import random
import logging
import httpx
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.config import settings
from app.services.auth_service import hash_password
from app.models.user import User, UserRole, PositionType
from app.models.organization import Zone, Team

logger = logging.getLogger("battle100")

async def init_zones_and_teams(db: AsyncSession):
    """
    初始化战区和战队字典数据
    """
    # 1. 检查并建立三大战区
    zone_result = await db.execute(select(Zone))
    zones = zone_result.scalars().all()
    zone_dict = {z.name: z for z in zones}

    zone_names = ["第一战区", "第二战区", "第三战区"]
    for idx, name in enumerate(zone_names):
        if name not in zone_dict:
            new_zone = Zone(name=name, sort_order=idx + 1)
            db.add(new_zone)
            await db.flush()
            zone_dict[name] = new_zone

    # 2. 检查并建立九大冲刺战队
    team_result = await db.execute(select(Team))
    teams = team_result.scalars().all()
    team_names = {t.name for t in teams}

    zone_team_mapping = {
        "第一战区": ["清远战队", "广州一战队", "广州二战队"],
        "第二战区": ["广州三战队（大数据）", "佛山战队", "湛江战队"],
        "第三战区": ["云浮战队", "东莞战队", "茂名战队"]
    }

    for zone_name, t_names in zone_team_mapping.items():
        z = zone_dict[zone_name]
        for t_name in t_names:
            if t_name not in team_names:
                new_team = Team(name=t_name, zone_id=z.id)
                db.add(new_team)
                await db.flush()

def map_dingtalk_dept_to_team_name(dept_name: str, parent_dept_names: list[str]) -> str | None:
    """
    根据钉钉的部门名称（及父部门名称）精准推断百日奋战战队名称。
    """
    if not dept_name:
        return None
        
    full_path = " / ".join(parent_dept_names + [dept_name]).lower()
    dept_lower = dept_name.lower()
    
    # 优先匹配其他分公司，避免分公司的技术部门被错误归入广州一/二战队
    if "清远" in full_path:
        return "清远战队"
    elif "佛山" in full_path:
        return "佛山战队"
    elif "湛江" in full_path:
        return "湛江战队"
    elif "云浮" in full_path:
        return "云浮战队"
    elif "东莞" in full_path:
        return "东莞战队"
    elif "茂名" in full_path:
        return "茂名战队"
        
    # 广州三战队 (大数据)
    if "大数据" in full_path:
        return "广州三战队（大数据）"
        
    # 广州一战队 (蓝色框：技术1巴、2巴、3巴, 5巴)
    if any(x in dept_lower for x in ["技术1巴", "技术一巴", "技术2巴", "技术二巴", "技术3巴", "技术三巴", "技术5巴", "技术五巴"]):
        return "广州一战队"
        
    # 广州二战队 (红色框：技术4巴、6巴、7巴、8巴)
    if any(x in dept_lower for x in ["技术4巴", "技术四巴", "技术6巴", "技术六巴", "技术7巴", "技术七巴", "技术8巴", "技术八巴"]):
        return "广州二战队"
        
    # 兜底匹配
    if "一部" in dept_lower or "一部门" in dept_lower:
        return "广州一战队"
    elif "二部" in dept_lower or "二部门" in dept_lower:
        return "广州二战队"
        
    return None

async def get_dingtalk_access_token() -> str:
    url = f"https://oapi.dingtalk.com/gettoken?appkey={settings.DINGTALK_APP_KEY}&appsecret={settings.DINGTALK_APP_SECRET}"
    async with httpx.AsyncClient() as client:
        resp = await client.get(url)
        data = resp.json()
        if data.get("errcode") == 0:
            return data.get("access_token")
        else:
            raise HTTPException(status_code=500, detail=f"获取钉钉 token 失败: {data}")

@router.post("/users/sync-dingtalk", summary="从钉钉通讯录同步活跃用户")
async def sync_users_from_dingtalk(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
):
    """
    通过钉钉 OpenAPI 同步全公司员工名单。
    获取真实岗位、真实手机号，并清除已离职员工。
    """
    await init_zones_and_teams(db)
    
    # 获取战队映射缓存
    team_res = await db.execute(select(Team))
    all_teams = team_res.scalars().all()
    team_by_name = {t.name: t.id for t in all_teams}
    team_ids = list(team_by_name.values())

    dingtalk_users = []
    is_mocked = False
    
    try:
        token = await get_dingtalk_access_token()
        
        async with httpx.AsyncClient(timeout=120) as client:
            # 1. 递归获取所有部门
            depts_to_process = [1]
            all_depts = {1: {"name": "根部门", "parent_id": None}}
            
            head = 0
            while head < len(depts_to_process):
                current_dept_id = depts_to_process[head]
                head += 1
                
                resp = await client.post(f"https://oapi.dingtalk.com/topapi/v2/department/listsub?access_token={token}", json={"dept_id": current_dept_id})
                data = resp.json()
                if data.get("errcode") == 0 and data.get("result"):
                    for child in data["result"]:
                        dept_id = child["dept_id"]
                        all_depts[dept_id] = {"name": child["name"], "parent_id": current_dept_id}
                        depts_to_process.append(dept_id)
            
            # ===== 白名单过滤：只导入红框内组织的人员 =====
            # 这些是根部门（1）的直接子部门中需要导入的
            ALLOWED_TOP_DEPTS = {
                "领导层", "财务部", "经营管理部", "商务办", "市场部",
                "技术中心", "幸福委员会", "行政部", "人力资源部", "投标部",
                "大数据部", "中地研究院",
                "广州分公司", "佛山分公司", "清远分公司",
                "湛江分公司", "云浮分公司", "东莞分公司", "茂名分公司"
            }
            
            # 找到白名单中顶层部门的 dept_id
            allowed_top_ids = set()
            for dept_id, info in all_depts.items():
                if info["parent_id"] == 1 and info["name"] in ALLOWED_TOP_DEPTS:
                    allowed_top_ids.add(dept_id)
            
            # 构建每个部门到顶层父部门的映射，判断是否在白名单分支下
            def is_dept_allowed(dept_id):
                """判断某个部门是否属于白名单分支"""
                if dept_id == 1:
                    return False  # 根部门本身不算
                curr = dept_id
                while curr and curr in all_depts:
                    if curr in allowed_top_ids:
                        return True
                    curr = all_depts[curr]["parent_id"]
                return False
            
            # 过滤出允许的部门列表
            allowed_depts = [d for d in depts_to_process if is_dept_allowed(d)]
            logger.info(f"白名单过滤：从 {len(depts_to_process)} 个部门中筛选出 {len(allowed_depts)} 个有效部门")
            
            # 2. 遍历白名单部门获取用户
            seen_userids = set()
            
            for dept_id in allowed_depts:
                cursor = 0
                while True:
                    resp = await client.post(f"https://oapi.dingtalk.com/topapi/v2/user/list?access_token={token}", json={
                        "dept_id": dept_id,
                        "cursor": cursor,
                        "size": 100
                    })
                    data = resp.json()
                    if data.get("errcode") != 0:
                        logger.warning(f"获取部门 {dept_id} 员工失败: {data}")
                        break
                        
                    page_result = data.get("result", {})
                    user_list = page_result.get("list", [])
                    
                    for u in user_list:
                        if u["userid"] not in seen_userids:
                            seen_userids.add(u["userid"])
                            
                            # 构建部门路径
                            path = []
                            curr = dept_id
                            while curr and curr in all_depts:
                                path.insert(0, all_depts[curr]["name"])
                                curr = all_depts[curr]["parent_id"]
                                
                            dingtalk_users.append({
                                "userid": u["userid"],
                                "name": u["name"],
                                "mobile": u.get("mobile", ""),
                                "title": u.get("title", ""),
                                "dept_name": all_depts[dept_id]["name"],
                                "dept_path": path
                            })
                            
                    if not page_result.get("has_more"):
                        break
                    cursor = page_result.get("next_cursor")
                    
    except Exception as e:
        logger.error(f"调用钉钉接口失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"钉钉接口调用失败，请检查网络或配置: {str(e)}")

    crm_users = {}
    try:
        conn = pymysql.connect(
            host=settings.CRM_DB_HOST,
            port=settings.CRM_DB_PORT,
            user=settings.CRM_DB_USER,
            password=settings.CRM_DB_PASSWORD,
            database=settings.CRM_DB_NAME,
            charset='utf8mb4',
            connect_timeout=3
        )
        cur = conn.cursor(pymysql.cursors.DictCursor)
        cur.execute("SELECT user_code, user_name, mobile, phone FROM js_sys_user WHERE status = '0'")
        rows = cur.fetchall()
        for r in rows:
            user_name = r["user_name"]
            mobile = r["mobile"]
            phone = r["phone"]
            user_code = r["user_code"]
            if mobile:
                crm_users[str(mobile).strip()] = user_code
            if phone:
                crm_users[str(phone).strip()] = user_code
            if user_name:
                crm_users[f"name_{user_name.strip()}"] = user_code
        cur.close()
        conn.close()
    except Exception as crm_err:
        logger.warning(f"同步钉钉用户时，连接 CRM 失败: {crm_err}")

    try:
        sync_count = 0
        skip_count = 0

        for u in dingtalk_users:
            name = u['name']
            phone = u['mobile']
            if not phone:
                skip_count += 1
                continue

            phone = str(phone).strip()
            name = str(name).strip()

            result = await db.execute(select(User).where(User.phone == phone))
            existing_user = result.scalar_one_or_none()

            # 提取钉钉真实岗位，推导岗位类别与战队
            position = u['title'] or ""
            dept_full_path = " / ".join(u['dept_path']) + " / " + u['dept_name']
            dept_lower = dept_full_path.lower()
            title_lower = position.lower()

            pos_type = PositionType.DELIVERY
            team_id = None
            
            # --- 1. 后台与中台判定（无战队） ---
            if "领导层" in dept_lower:
                pos_type = PositionType.BACK_OFFICE
                team_id = None
            elif any(x in dept_lower for x in ["财务", "经营管理", "商务办", "市场", "技术中心", "幸福", "行政", "人力", "投标"]):
                pos_type = PositionType.MIDDLE_OFFICE
                team_id = None
            else:
                # --- 分配战队逻辑 ---
                team_name = map_dingtalk_dept_to_team_name(u['dept_name'], u['dept_path'])
                
                # 特殊兜底处理：广州交付综管巴 默认划入广州二战队
                if "广州" in dept_lower and ("综管" in dept_lower or "综合管理" in dept_lower) and not team_name:
                    team_name = "广州二战队"
                    
                if team_name:
                    team_id = team_by_name.get(team_name)
                else:
                    team_id = None  # 未匹配到战队的员工不分配
                    
                # --- 2. 业务线的岗位判定 ---
                if any(x in u['dept_name'].lower() for x in ["综合管理", "综管"]):
                    pos_type = PositionType.MANAGEMENT
                elif "主管" in title_lower:
                    pos_type = PositionType.MANAGEMENT
                elif "营销" in dept_lower or "销售" in title_lower:
                    pos_type = PositionType.MARKETING
                elif "技术" in dept_lower or "技术" in title_lower:
                    pos_type = PositionType.TECHNICAL
                else:
                    pos_type = PositionType.DELIVERY

            # 解析三级巴：排除第一个"根部门"后，长度达到或超过3级，则第3级为三级巴
            clean_path = u['dept_path'][1:] if u['dept_path'] and u['dept_path'][0] == "根部门" else u['dept_path']
            third_class_bar = clean_path[2] if len(clean_path) >= 3 else None

            # 匹配 CRM 用户 ID
            crm_user_id = None
            if phone and phone in crm_users:
                crm_user_id = crm_users[phone]
            elif name and f"name_{name}" in crm_users:
                crm_user_id = crm_users[f"name_{name}"]

            # 强制覆盖 6 位营销骨干与巴长战队/岗位，防止被钉钉中后台部门判定覆盖
            forced_teams = {
                "王浩亮": ("广州一战队", PositionType.MARKETING),
                "朱海": ("广州一战队", PositionType.MARKETING),
                "陈浩龙": ("广州一战队", PositionType.MARKETING),
                "张永雄": ("广州二战队", PositionType.MARKETING),
                "李健鹏": ("广州二战队", PositionType.MARKETING),
                "刘训东": ("茂名战队", PositionType.MARKETING)
            }
            if name in forced_teams:
                forced_t_name, forced_pos = forced_teams[name]
                team_id = team_by_name.get(forced_t_name)
                pos_type = forced_pos

            if existing_user:
                existing_user.name = name
                existing_user.position = position if position else existing_user.position
                existing_user.position_type = pos_type
                existing_user.third_class_bar = third_class_bar
                existing_user.team_id = team_id
                existing_user.dingtalk_id = u['userid']
                if crm_user_id:
                    existing_user.crm_user_id = crm_user_id
                db.add(existing_user)
            else:
                new_user = User(
                    name=name,
                    phone=phone,
                    password_hash=hash_password("123456"),
                    position=position,
                    position_type=pos_type,
                    third_class_bar=third_class_bar,
                    role=UserRole.STAFF,
                    team_id=team_id,
                    dingtalk_id=u['userid'],
                    crm_user_id=crm_user_id,
                    is_active=True
                )
                db.add(new_user)

            sync_count += 1

        await db.flush()

        return {
            "message": f"钉钉通讯录同步成功！共处理 {sync_count} 个员工，跳过 {skip_count} 个无手机号记录。" + (" (演示 Mock 模式)" if is_mocked else ""),
            "sync_count": sync_count,
            "skip_count": skip_count,
            "mode": "mock" if is_mocked else "real"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"同步处理失败: {str(e)}")


@router.post("/goals/sync-crm", summary="从CRM数据库同步有效线索")
async def sync_goals_from_crm(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.TARGET_OFFICER)),
):
    """
    直连 CRM 数据库抓取所有处于 25%-75% 推进阶段且未中止的线索。
    若连接超时或失败，则降级为 Mock 模式，随机生成 30-60 条符合业务进度的有效线索。
    """
    conn = None
    crm_leads = []
    is_mocked = False

    try:
        conn = pymysql.connect(
            host=settings.CRM_DB_HOST,
            port=settings.CRM_DB_PORT,
            user=settings.CRM_DB_USER,
            password=settings.CRM_DB_PASSWORD,
            database=settings.CRM_DB_NAME,
            charset='utf8mb4',
            connect_timeout=5
        )
        cur = conn.cursor(pymysql.cursors.DictCursor)
        # 提取 25%-75% 未中止线索
        cur.execute("""
            SELECT id, name, customer_name, expect_money, progress, market_user_id, feed_back_date 
            FROM zdcrm_business_opportunity 
            WHERE progress BETWEEN 25 AND 75 
              AND (is_suspension = '0' OR is_suspension IS NULL)
        """)
        crm_leads = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"无法直连 CRM 获取线索: {str(e)}")
        raise HTTPException(status_code=500, detail=f"无法连接到 CRM 数据库，请检查网络或配置: {str(e)}")

    try:
        # 对齐数据格式以供前端直接展示或进一步指标绑定
        formatted_leads = []
        for lead in crm_leads:
            formatted_leads.append({
                "id": str(lead["id"]),
                "name": lead["name"],
                "customer_name": lead["customer_name"],
                "expect_amount": float(lead["expect_money"]) if lead["expect_money"] else 0.0,
                "progress": int(lead["progress"]) if lead["progress"] else 0,
                "owner_id": lead["market_user_id"]
            })

        return {
            "message": f"成功同步 CRM 里的 {len(formatted_leads)} 条有效推进线索。" + (" (降级 Mock 模式)" if is_mocked else ""),
            "leads_count": len(formatted_leads),
            "leads": formatted_leads,
            "mode": "mock" if is_mocked else "real"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"同步线索数据失败: {str(e)}")


@router.get("/goals/personal/export", summary="导出个人奋斗目标与完成情况Excel")
async def export_personal_goals(
    export_type: str = Query("goals", description="导出类型: goals(奋斗目标) 或 actuals(完成情况)"),
    keyword: str | None = Query(None, description="姓名或手机号模糊搜索"),
    team_id: int | None = Query(None, description="战队ID筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 1. 查询符合条件的所有个人目标和用户信息
    stmt = select(PersonalGoal, User).join(User, PersonalGoal.user_id == User.id)
    
    if keyword:
        keyword_filter = or_(User.name.contains(keyword), User.phone.contains(keyword))
        stmt = stmt.where(keyword_filter)
        
    if team_id:
        stmt = stmt.where(User.team_id == team_id)
        
    stmt = stmt.order_by(User.id, PersonalGoal.id)
    results = await db.execute(stmt)
    rows = results.all()
    
    # 2. 转换战队映射
    team_res = await db.execute(select(Team))
    team_map = {t.id: t.name for t in team_res.scalars().all()}
    
    # 3. 按用户做聚合（Pivot）
    user_goals = {}
    user_info = {}
    for goal, user in rows:
        uid = user.id
        if uid not in user_goals:
            user_goals[uid] = {}
            user_info[uid] = user
        user_goals[uid][goal.goal_type.value if hasattr(goal.goal_type, 'value') else goal.goal_type] = goal

    # 如果是实际完成情况，则需要计算系统实际值
    system_actuals = {}
    if export_type == "actuals" and user_goals:
        from app.api.goals import fetch_users_system_actual_values
        system_actuals = await fetch_users_system_actual_values(db, list(user_goals.keys()))

    # 4. 用 openpyxl 写入 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "个人实际完成情况" if export_type == "actuals" else "个人奋斗目标"
    
    from app.models.goal import GoalType
    kpis = [
        (GoalType.CONTRACT_AMOUNT.value, "新签/续签合同额", "万元"),
        (GoalType.HAPPINESS_ACTION.value, "客户幸福动作完成数", "次"),
        (GoalType.TRIANGLE_COUNT.value, "售前铁三角联动次数", "次"),
        (GoalType.LEADS_COUNT.value, "有效线索数", "条"),
        (GoalType.LEADS_CONVERSION_RATE.value, "线索转化率", "%"),
        (GoalType.NEW_CUSTOMER_COUNT.value, "新客户数", "个"),
        (GoalType.HAPPINESS_STORY_COUNT.value, "幸福故事数", "个"),
        (GoalType.CONTRACT_COUNT.value, "新签合同单数", "个")
    ]
    
    # 写入表头
    if export_type == "goals":
        headers = ["姓名", "岗位", "手机号", "战区", "归属战队"]
        for _, label, unit in kpis:
            headers.extend([f"{label}({unit}) - 基础", f"{label}({unit}) - 挑战"])
        ws.append(headers)
    else:
        # 完成情况：实际/目标
        headers = ["姓名", "岗位", "手机号", "战区", "归属战队"]
        for _, label, unit in kpis:
            headers.append(f"{label}(实际/基础){unit}")
        ws.append(headers)
        
    def get_zone_name_by_team_id(tid: int | None) -> str:
        if not tid:
            return "未分配"
        if tid in [1, 2, 3]:
            return "第一战区"
        if tid in [4, 5, 6]:
            return "第二战区"
        if tid in [7, 8, 9]:
            return "第三战区"
        return "未分配"

    # 写入数据行
    for uid, goals in user_goals.items():
        user = user_info[uid]
        team_name = team_map.get(user.team_id, "未分配") if user.team_id else "未分配"
        zone_name = get_zone_name_by_team_id(user.team_id)
        
        row_data = [user.name, user.position or "—", user.phone, zone_name, team_name]
        
        for kpi_key, _, _ in kpis:
            goal = goals.get(kpi_key)
            if export_type == "goals":
                if goal:
                    row_data.extend([goal.base_target, goal.challenge_target])
                else:
                    row_data.extend(["—", "—"])
            else:
                # 实际完成值 / 基础目标值
                if not goal and kpi_key != GoalType.CONTRACT_AMOUNT.value:
                    row_data.append("—")
                else:
                    user_sys_vals = system_actuals.get(uid, {})
                    sys_val = user_sys_vals.get(kpi_key, 0.0)
                    actual_val = goal.actual_value if (goal and goal.actual_value is not None) else sys_val
                    base_target = goal.base_target if goal else 0.0
                    
                    # 格式为： 实际 / 目标
                    row_data.append(f"{actual_val} / {base_target}")
                    
        ws.append(row_data)
        
    # 保存并返回二进制文件流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "personal_actuals_export.xlsx" if export_type == "actuals" else "personal_goals_export.xlsx"
    
    # 记录审计日志
    await log_action(
        db=db,
        user=current_user,
        action_type="EXPORT",
        target_module="goal",
        target_id="0",
        description=f"导出了个人目标数据，类型: {export_type}",
        before_state=None,
        after_state=None,
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


