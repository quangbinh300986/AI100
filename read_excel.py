# -*- coding: utf-8 -*-
import openpyxl
import json

wb = openpyxl.load_workbook(r'百日奋战目标、定义及各项工作安排.xlsx')

print("=" * 60)
print("工作表名称列表:")
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"  [{i}] {name} (行数: {ws.max_row}, 列数: {ws.max_column})")

print("\n" + "=" * 60)

# 遍历每个工作表，打印前几行内容
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"工作表: {name}")
    print(f"{'='*60}")
    
    max_rows_to_print = min(ws.max_row, 30)  # 每个sheet最多打印30行
    for row in ws.iter_rows(min_row=1, max_row=max_rows_to_print, values_only=False):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"[{cell.column_letter}{cell.row}]{val}")
        if row_data:
            print(" | ".join(row_data))
